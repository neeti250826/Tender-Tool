#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import html
import json
import logging
import os
import re
import sys
import time
import unicodedata
import zipfile
from urllib.parse import unquote, urljoin, urlparse
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Dict, Iterable, Iterator, List, Optional, Sequence, Tuple

import requests
from requests import RequestException

API_URL = "https://ezamowienia.gov.pl/mo-board/api/v1/notice"
NOTICE_DETAIL_URL = "https://ezamowienia.gov.pl/mo-client-board/bzp/notice-details/id/{object_id}"
PROCEDURE_URLS = (
    "https://ezamowienia.gov.pl/mp-client/search/list/{tender_id}",
    "https://ezamowienia.gov.pl/mp-client/tenders/{tender_id}",
)

DEFAULT_NOTICE_TYPES = ["ContractNotice"]
ALLOWED_NOTICE_TYPES = ["ContractNotice", "TenderResultNotice", "ContractPerformingNotice"]
MEDICAL_CPV_PREFIXES = ("331", "336", "851")
MEDICAL_PATTERNS = (
    r"\bszpital",
    r"\bmedycz",
    r"\bzdrow",
    r"\bklinicz",
    r"\bklinika",
    r"\bokulist",
    r"\bpsychiatr",
    r"\bapte",
    r"\busg\b",
    r"\blaborator",
    r"\bdiagnost",
    r"\brehabilit",
    r"\bambulans",
    r"\bambulatoryj",
    r"\bpieleg",
    r"\bratownict",
    r"\bkrwiodaw",
    r"\bkrwiolecz",
    r"\bodczynnik",
    r"\blecznic",
    r"\bfarmaceut",
    r"\bstomatolog",
    r"\btransportu sanitarnego",
    r"\buslug medyczn",
    r"\bopieki zdrowot",
)

# Text fallback for medical filtering. CPV is the primary signal. Avoid broad
# words like "zdrow" because construction notices often contain generic
# occupational health/safety phrases.
STRICT_MEDICAL_TEXT_PATTERNS = (
    r"\bszpital", r"\bmedycz", r"\bklinicz", r"\bklinika", r"\bokulist",
    r"\bpsychiatr", r"\bapte", r"\busg\b", r"\blaborator", r"\bdiagnost",
    r"\brehabilit", r"\bambulans", r"\bambulatoryj", r"\bpieleg",
    r"\bratownict", r"\bkrwiodaw", r"\bkrwiolecz", r"\bodczynnik",
    r"\blecznic", r"\bfarmaceut", r"\bstomatolog",
    r"\btransportu sanitarnego", r"\buslug medyczn", r"\bopieki zdrowot",
)

NON_MEDICAL_CPV_PREFIXES = (
    "03", "15", "16", "18", "19", "22", "24", "30", "31", "32", "34",
    "35", "37", "38", "39", "41", "42", "43", "44", "45", "48", "50",
    "55", "60", "63", "64", "65", "66", "70", "71", "72", "73", "75",
    "77", "79", "80", "90", "92", "98",
)

ORIGINAL_NOTICE_LOOKBACK_DAYS = 730
ITEM_DOCUMENT_NAME_PATTERNS = (
    r"formularz",
    r"asortyment",
    r"cenow",
    r"opis",
    r"opz",
    r"swz",
    r"zalacznik",
    r"załącznik",
    r"pakiet",
    r"przedmiot",
    r"specyfikacj",
    r"wycena",
    r"ilo[sś]c",
)

NA = "N/A"

BASE_FIELDS = [
    "source",
    "country",
    "country_cc",
    "publication_date",
    "closing_date",
    "title",
    "description",
    "buyer",
    "classification",
    "status",
    "currency",
    "amount",
    "awarding_supplier_name",
    "awarded_currency",
    "awarded_value",
    "awarded_date",
    "contract_period",
    "item_no",
    "item_desc",
    "item_uom",
    "item_quantity",
    "item_unit_price",
    "item_award",
    "notice_id",
    "notice_url",
    "tender_id",
    "notice_type",
    "query_text",
    "scraped_at",
    "dedup_key",
]

TRANSLATE_SOURCE_FIELDS = [
    "title",
    "description",
    "buyer",
    "classification",
    "status",
    "item_desc",
    "item_uom",
    "awarding_supplier_name",
]

TRANSLATION_FIELDS = [f"{field}_en" for field in TRANSLATE_SOURCE_FIELDS]


def build_fieldnames() -> List[str]:
    """Put English columns directly next to their Polish source columns."""
    ordered: List[str] = []
    for field in BASE_FIELDS:
        ordered.append(field)
        if field in TRANSLATE_SOURCE_FIELDS:
            ordered.append(f"{field}_en")
    for field in TRANSLATION_FIELDS:
        if field not in ordered:
            ordered.append(field)
    return ordered


FIELDNAMES = build_fieldnames()

UNIT_PATTERN = (
    r"opak(?:\.|owanie|owania|owan)?|opk\.?|op\.?|szt\.?|sztuk(?:a|i)?|"
    r"kpl\.?|kompl\.?|komplet(?:y)?|zestaw(?:y)?|para|par|amp\.?|ampul(?:ka|ki|ek)?|"
    r"fiol\.?|fiolk(?:a|i|ek)?|tabl\.?|tablet(?:ka|ki|ek)?|kaps\.?|kapsulk(?:a|i|ek)?|"
    r"sasz(?:etka|etki)?|wstrzykiwacz(?:y|e)?|ampulkostrzykawk(?:a|i|ek)?|"
    r"kg|g|mg|mcg|ug|ml|l|m2|m3|mb"
)

POLISH_TRANS = str.maketrans(
    {
        "\u0105": "a",
        "\u0107": "c",
        "\u0119": "e",
        "\u0142": "l",
        "\u0144": "n",
        "\u00f3": "o",
        "\u015b": "s",
        "\u017a": "z",
        "\u017c": "z",
        "\u0104": "A",
        "\u0106": "C",
        "\u0118": "E",
        "\u0141": "L",
        "\u0143": "N",
        "\u00d3": "O",
        "\u015a": "S",
        "\u0179": "Z",
        "\u017b": "Z",
    }
)


def normalize_text(value: Any, keep_newlines: bool = False) -> str:
    if value is None:
        return ""
    if isinstance(value, (list, tuple, set)):
        value = " ".join(str(part) for part in value if part is not None)
    text = html.unescape(str(value))
    text = re.sub(r"(?i)<\s*br\s*/?\s*>", "\n", text)
    text = re.sub(r"(?i)</\s*(p|div|h\d|li|tr|table|section)\s*>", "\n", text)
    text = re.sub(r"<[^>]+>", " ", text)
    text = text.replace("\xa0", " ").replace("\u200b", " ")
    text = text.replace("\r", "\n").replace("\t", " ")
    text = re.sub(r"[ \f\v]+", " ", text)
    if keep_newlines:
        text = re.sub(r" *\n+ *", "\n", text)
    else:
        text = re.sub(r"\s+", " ", text)
    return text.strip()


def one_line(value: Any) -> str:
    return normalize_text(value, keep_newlines=False)


def fold_text(value: Any) -> str:
    text = normalize_text(value, keep_newlines=False).translate(POLISH_TRANS)
    text = unicodedata.normalize("NFKD", text)
    text = text.encode("ascii", "ignore").decode("ascii")
    return text.lower()


def nonblank(value: Any, default: str = NA) -> str:
    text = one_line(value)
    return text if text else default


def parse_iso_date(value: str) -> date:
    return datetime.strptime(value, "%Y-%m-%d").date()


def parse_any_date(value: Any) -> Optional[date]:
    """Parse common Polish/API date formats and return a date object.

    Supported examples: 2024-01-31, 2024-01-31T10:00:00Z,
    31.01.2024, 31/01/2024, and 31-01-2024.
    """
    text = one_line(value)
    if not text or text == NA:
        return None

    iso_match = re.search(r"(?<!\d)(\d{4})-(\d{2})-(\d{2})(?!\d)", text)
    if iso_match:
        try:
            return date(int(iso_match.group(1)), int(iso_match.group(2)), int(iso_match.group(3)))
        except ValueError:
            pass

    dmy_match = re.search(r"(?<!\d)(\d{1,2})[./-](\d{1,2})[./-](\d{4})(?!\d)", text)
    if dmy_match:
        try:
            return date(int(dmy_match.group(3)), int(dmy_match.group(2)), int(dmy_match.group(1)))
        except ValueError:
            pass
    return None


def date_only(value: Any) -> str:
    parsed = parse_any_date(value)
    if parsed:
        return parsed.isoformat()
    text = one_line(value)
    return text if text else ""


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def chunk_dates(start: date, end: date, window_days: int) -> Iterator[Tuple[date, date]]:
    current = start
    step = max(1, int(window_days))
    while current <= end:
        window_end = min(end, current + timedelta(days=step - 1))
        yield current, window_end
        current = window_end + timedelta(days=1)


def request_json_with_retries(
    session: requests.Session,
    params: Dict[str, Any],
    retries: int = 3,
    retry_sleep: float = 2.0,
) -> List[Dict[str, Any]]:
    last_error: Optional[BaseException] = None
    for attempt in range(1, retries + 1):
        try:
            response = session.get(API_URL, params=params, timeout=90)
            response.raise_for_status()
            payload = response.json()
            if isinstance(payload, list):
                return payload
            if isinstance(payload, dict):
                for key in ("items", "data", "results", "notices"):
                    value = payload.get(key)
                    if isinstance(value, list):
                        return value
            raise ValueError(f"Unexpected API response format: {type(payload).__name__}")
        except (RequestException, ValueError) as exc:
            last_error = exc
            logging.warning("API request failed on attempt %s/%s: %s", attempt, retries, exc)
            if attempt < retries:
                time.sleep(retry_sleep * attempt)
    raise RuntimeError(f"API request failed after {retries} attempts: {last_error}")


def iter_notices(
    session: requests.Session,
    notice_type: str,
    date_from: str,
    date_to: str,
    page_size: int,
    window_days: int,
    request_sleep: float,
) -> Iterator[Tuple[str, str, int, List[Dict[str, Any]]]]:
    start = parse_iso_date(date_from)
    end = parse_iso_date(date_to)
    for chunk_start, chunk_end in chunk_dates(start, end, window_days):
        page_number = 1
        page_signatures = set()
        while True:
            params = {
                "PageNumber": page_number,
                "PageSize": page_size,
                "NoticeType": notice_type,
                "PublicationDateFrom": chunk_start.isoformat(),
                "PublicationDateTo": chunk_end.isoformat(),
            }
            notices = request_json_with_retries(session, params)
            signature = tuple(
                one_line(notice.get("noticeNumber")) or one_line(notice.get("tenderId")) or one_line(notice.get("objectId"))
                for notice in notices[:5]
            )
            if signature and signature in page_signatures:
                logging.warning(
                    "API returned a repeated page for %s %s..%s page %s; stopping this window.",
                    notice_type,
                    chunk_start,
                    chunk_end,
                    page_number,
                )
                break
            page_signatures.add(signature)
            yield chunk_start.isoformat(), chunk_end.isoformat(), page_number, notices
            if len(notices) < page_size:
                break
            page_number += 1
            if request_sleep:
                time.sleep(request_sleep)


def get_html_text(notice: Dict[str, Any]) -> str:
    return normalize_text(notice.get("htmlBody"), keep_newlines=True)


def heading_id_pattern(section_id: str) -> re.Pattern[str]:
    escaped = re.escape(section_id)
    return re.compile(rf"^\s*#*\s*{escaped}(?:\.|\)|\s)", re.I)


def line_matches_section(line: str, section_id: str) -> bool:
    clean = line.strip().lstrip("#").strip()
    return bool(heading_id_pattern(section_id).search(clean))


def is_numbered_heading(line: str) -> bool:
    stripped = line.strip()
    return (
        bool(re.match(r"^\s*#{1,6}\s+\d+(?:\.\d+)*[.)]", stripped))
        or bool(re.match(r"^\s*\d+(?:\.\d+)+[.)]", stripped))
        or bool(re.match(r"^\s*##\s+SEKCJA\b", stripped, re.I))
    )


def strip_heading_value(line: str, section_id: str) -> str:
    clean = line.strip().lstrip("#").strip()
    clean = re.sub(rf"^\s*{re.escape(section_id)}(?:\.|\))?\s*", "", clean, flags=re.I)
    if ":" in clean:
        return clean.split(":", 1)[1].strip()
    return ""


def extract_numbered_sections(text: str, section_ids: Sequence[str]) -> List[str]:
    lines = normalize_text(text, keep_newlines=True).splitlines()
    sections: List[str] = []
    for index, line in enumerate(lines):
        matched_id = next((sid for sid in section_ids if line_matches_section(line, sid)), None)
        if not matched_id:
            continue
        body: List[str] = []
        inline_value = strip_heading_value(line, matched_id)
        if inline_value:
            body.append(inline_value)
        for next_line in lines[index + 1 :]:
            if is_numbered_heading(next_line):
                break
            if next_line.strip():
                body.append(next_line)
        section = normalize_text("\n".join(body), keep_newlines=False)
        if section:
            sections.append(section)
    return sections


def first_section(text: str, section_ids: Sequence[str]) -> str:
    sections = extract_numbered_sections(text, section_ids)
    return sections[0] if sections else ""


def extract_description(html_text: str, title: str) -> str:
    for section_ids in (("4.2.2",), ("4.5.1",), ("3.8",), ("2.3",)):
        section = first_section(html_text, section_ids)
        if section:
            return section
    return one_line(title)


def extract_cpv_from_text(text: str) -> str:
    matches = re.findall(r"\b\d{8}-\d\b(?:\s*\([^)]+\))?", normalize_text(text, keep_newlines=False))
    seen: List[str] = []
    for match in matches:
        cleaned = one_line(match)
        if cleaned and cleaned not in seen:
            seen.append(cleaned)
    return ",".join(seen)


def extract_labeled_line(text: str, label_fragments: Sequence[str]) -> str:
    lines = normalize_text(text, keep_newlines=True).splitlines()
    folded_fragments = [fold_text(fragment) for fragment in label_fragments]
    for index, line in enumerate(lines):
        folded = fold_text(line)
        if any(fragment in folded for fragment in folded_fragments):
            if ":" in line:
                value = line.split(":", 1)[1].strip()
                if value:
                    return one_line(value)
            for next_line in lines[index + 1 : index + 4]:
                if next_line.strip() and not is_numbered_heading(next_line):
                    return one_line(next_line)
    return ""


def extract_labeled_context(text: str, label_fragments: Sequence[str], following_lines: int = 4) -> str:
    """Return the matching label line plus a few following lines for date/duration parsing."""
    lines = normalize_text(text, keep_newlines=True).splitlines()
    folded_fragments = [fold_text(fragment) for fragment in label_fragments]
    contexts: List[str] = []
    for index, line in enumerate(lines):
        folded = fold_text(line)
        if any(fragment in folded for fragment in folded_fragments):
            contexts.extend(lines[index : index + following_lines + 1])
    return one_line(" ".join(contexts))


def extract_date_after_labels(text: str, label_fragments: Sequence[str], following_lines: int = 4) -> str:
    context = extract_labeled_context(text, label_fragments, following_lines=following_lines)
    return date_only(context)


def days_between(start_value: Any, end_value: Any, inclusive: bool = False) -> str:
    start_date = parse_any_date(start_value)
    end_date = parse_any_date(end_value)
    if not start_date or not end_date:
        return ""
    delta = (end_date - start_date).days
    if delta < 0:
        return ""
    if inclusive:
        delta += 1
    return str(delta)


def extract_duration_days(text: str) -> str:
    """Extract a duration and convert it to days.

    Exact day values are preferred. Month/year values are converted using
    procurement-friendly calendar approximations: 30 days/month and 365 days/year.
    """
    folded = fold_text(text)
    if not folded:
        return ""
    patterns = (
        (r"(?P<num>\d+(?:[,.]\d+)?)\s*(?:dni|dzien|dnia|days?)\b", Decimal("1")),
        (r"(?P<num>\d+(?:[,.]\d+)?)\s*(?:tygodni|tydzien|tygodnie|weeks?)\b", Decimal("7")),
        (r"(?P<num>\d+(?:[,.]\d+)?)\s*(?:miesiac|miesiace|miesiecy|m-cy|months?)\b", Decimal("30")),
        (r"(?P<num>\d+(?:[,.]\d+)?)\s*(?:rok|lata|lat|years?)\b", Decimal("365")),
    )
    for pattern, multiplier in patterns:
        match = re.search(pattern, folded, flags=re.I)
        if not match:
            continue
        try:
            value = Decimal(match.group("num").replace(",", ".")) * multiplier
        except InvalidOperation:
            continue
        return str(int(value.to_integral_value(rounding=ROUND_HALF_UP)))
    return ""


def extract_supplier(notice: Dict[str, Any], html_text: str) -> str:
    contractors = notice.get("contractors") or []
    if isinstance(contractors, list) and contractors:
        first = contractors[0] or {}
        if isinstance(first, dict):
            supplier = one_line(first.get("contractorName") or first.get("name"))
            if supplier:
                return supplier
    supplier = extract_labeled_line(
        html_text,
        (
            "nazwa wykonawcy",
            "dane wykonawcy",
            "wykonawca",
            "contractorName",
        ),
    )
    supplier = re.sub(r"^(nazwa|wykonawca)\s*[:\-]\s*", "", supplier, flags=re.I).strip()
    return supplier


def parse_decimal(value: Any) -> Optional[Decimal]:
    text = one_line(value).replace(" ", "").replace(",", ".")
    if not text or text == NA:
        return None
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def format_decimal(value: Decimal) -> str:
    quantized = value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    text = format(quantized, "f")
    return text.rstrip("0").rstrip(".") if "." in text else text


def find_money_after_labels(text: str, labels: Sequence[str]) -> Tuple[str, str]:
    lines = normalize_text(text, keep_newlines=True).splitlines()
    folded_labels = [fold_text(label) for label in labels]
    money_re = re.compile(r"(?P<amount>\d[\d\s.,]*\d|\d)\s*(?P<currency>PLN|EUR|USD|GBP|CHF)\b", re.I)
    for index, line in enumerate(lines):
        folded = fold_text(line)
        if any(label in folded for label in folded_labels):
            window = " ".join(lines[index : index + 4])
            match = money_re.search(window)
            if match:
                return one_line(match.group("currency")).upper(), one_line(match.group("amount")).replace(" ", "")
    return "", ""


def extract_money(notice_type: str, html_text: str) -> Tuple[str, str, str, str, str]:
    currency = ""
    amount = ""
    awarded_currency = ""
    awarded_value = ""
    item_award = ""

    if notice_type == "ContractNotice":
        currency, amount = find_money_after_labels(
            html_text,
            (
                "szacunkowa wartosc",
                "wartosc zamowienia",
                "calkowita wartosc zamowienia",
                "wartosc bez vat",
            ),
        )
    elif notice_type == "TenderResultNotice":
        awarded_currency, awarded_value = find_money_after_labels(
            html_text,
            (
                "cena wybranej oferty",
                "wartosc wybranej oferty",
                "wartosc umowy",
                "calkowita wartosc umowy",
            ),
        )
        currency = awarded_currency
        amount = awarded_value
        item_award = awarded_value
    elif notice_type == "ContractPerformingNotice":
        awarded_currency, awarded_value = find_money_after_labels(
            html_text,
            (
                "wartosc umowy",
                "cena wybranej oferty",
                "wartosc wykonanej umowy",
            ),
        )
        currency = awarded_currency
        amount = awarded_value
        item_award_currency, item_award_value = find_money_after_labels(
            html_text,
            (
                "wartosc wykonania",
                "laczna wartosc wynagrodzenia",
                "zmiana wartosci umowy",
            ),
        )
        if item_award_value:
            item_award = item_award_value
            if not awarded_currency:
                awarded_currency = item_award_currency
            if not currency:
                currency = item_award_currency
    return currency, amount, awarded_currency, awarded_value, item_award


AWARDED_DATE_LABELS = (
    "data wyboru najkorzystniejszej oferty",
    "data wyboru oferty",
    "data udzielenia zamowienia",
    "data udzielenia zamówienia",
    "data zawarcia umowy",
    "data zawarcia umowy w sprawie zamowienia publicznego",
    "data zawarcia umowy w sprawie zamówienia publicznego",
    "data rozstrzygniecia",
    "data rozstrzygnięcia",
)

CONTRACT_START_DATE_LABELS = (
    "data zawarcia umowy",
    "data rozpoczecia",
    "data rozpoczęcia",
    "data poczatkowa",
    "data początkowa",
    "termin rozpoczecia",
    "termin rozpoczęcia",
    "poczatek okresu realizacji",
    "początek okresu realizacji",
)

CONTRACT_END_DATE_LABELS = (
    "termin wykonania umowy",
    "termin realizacji umowy",
    "termin realizacji zamowienia",
    "termin realizacji zamówienia",
    "okres realizacji zamowienia",
    "okres realizacji zamówienia",
    "data zakonczenia",
    "data zakończenia",
    "data koncowa",
    "data końcowa",
)

CONTRACT_PERIOD_LABELS = (
    "okres realizacji zamowienia",
    "okres realizacji zamówienia",
    "okres obowiazywania umowy",
    "okres obowiązywania umowy",
    "termin wykonania umowy",
    "termin realizacji umowy",
    "termin realizacji zamowienia",
    "termin realizacji zamówienia",
    "czas trwania",
)


def extract_awarded_date(notice: Dict[str, Any], html_text: str, notice_type: str) -> str:
    """Return the best available award/contract date as YYYY-MM-DD."""
    direct_fields = (
        "awardDate",
        "awardedDate",
        "contractAwardDate",
        "contractDate",
        "contractConclusionDate",
        "agreementDate",
        "conclusionDate",
        "procedureResultDate",
    )
    for field in direct_fields:
        value = date_only(notice.get(field))
        if value:
            return value

    date_from_html = extract_date_after_labels(html_text, AWARDED_DATE_LABELS, following_lines=4)
    if date_from_html:
        return date_from_html

    # For completed-contract notices, the publication date is not the award date.
    # Keep it as a last fallback for result notices only when no award date exists in the body.
    if notice_type == "TenderResultNotice":
        return date_only(notice.get("publicationDate"))
    return ""


def extract_contract_period(html_text: str, notice_type: str, awarded_date: str = "", inclusive: bool = False) -> str:
    """Return contract period as a number of days, not a date range.

    The function first tries exact start/end dates. If those are unavailable,
    it parses explicit duration text such as "90 dni", "12 miesięcy", or "2 lata".
    """
    start = extract_date_after_labels(html_text, CONTRACT_START_DATE_LABELS, following_lines=4)
    end = extract_date_after_labels(html_text, CONTRACT_END_DATE_LABELS, following_lines=5)

    if not start and awarded_date:
        start = awarded_date

    exact_days = days_between(start, end, inclusive=inclusive)
    if exact_days:
        return exact_days

    label_context = extract_labeled_context(html_text, CONTRACT_PERIOD_LABELS, following_lines=6)
    section_context = " ".join(extract_numbered_sections(html_text, ("4.2.10", "4.2.11", "5.2", "4.2")))
    duration_days = extract_duration_days(" ".join([label_context, section_context]))
    if duration_days:
        return duration_days

    # Last resort: search only a small window after explicit period markers.
    # This avoids accidentally using unrelated values such as offer-validity days.
    folded_body = fold_text(html_text)
    for marker in ("okres realizacji", "termin wykonania umowy", "termin realizacji", "czas trwania"):
        marker_index = folded_body.find(marker)
        if marker_index >= 0:
            duration = extract_duration_days(folded_body[marker_index : marker_index + 500])
            if duration:
                return duration
    return ""


def cpv_codes_from_notice(notice: Dict[str, Any]) -> List[str]:
    text = " ".join(
        [
            one_line(notice.get("cpvCode")),
            extract_cpv_from_text(one_line(notice.get("htmlBody"))[:20000]),
        ]
    )
    codes = re.findall(r"\b\d{8}-\d\b|\b\d{8}\b", text)
    seen: List[str] = []
    for code in codes:
        normalized = code[:8]
        if normalized and normalized not in seen:
            seen.append(normalized)
    return seen


def is_medical_notice(notice: Dict[str, Any]) -> bool:
    """Return True for medical/healthcare procurements.

    CPV is the primary signal. Text is only a fallback. This prevents false
    positives like road/building tenders that mention generic health-and-safety
    text, while keeping real medical products/services.
    """
    cpv_codes = cpv_codes_from_notice(notice)
    if any(code.startswith(prefix) for code in cpv_codes for prefix in MEDICAL_CPV_PREFIXES):
        return True
    if cpv_codes and any(code.startswith(prefix) for code in cpv_codes for prefix in NON_MEDICAL_CPV_PREFIXES):
        return False

    text = " ".join(
        [
            one_line(notice.get("orderObject")),
            one_line(notice.get("organizationName")),
            one_line(notice.get("htmlBody"))[:8000],
        ]
    )
    folded = fold_text(text)
    folded = re.sub(r"\bochron[ayie]*\s+zdrowia\b", " ", folded)
    return any(re.search(pattern, folded, flags=re.I) for pattern in STRICT_MEDICAL_TEXT_PATTERNS)


def status_for_notice_type(notice_type: str) -> str:
    return {
        "ContractNotice": "Open",
        "TenderResultNotice": "Awarded",
        "ContractPerformingNotice": "Complete",
    }.get(notice_type, notice_type or "Unknown")


def extract_closing_date(notice: Dict[str, Any], html_text: str) -> str:
    direct = date_only(notice.get("submittingOffersDate"))
    if direct:
        return direct
    candidate = extract_labeled_line(html_text, ("termin skladania ofert", "termin skladania"))
    return date_only(candidate) or candidate


ITEM_CONTEXT_REJECT_PATTERNS = (
    r"\bokres realizacji\b",
    r"\bokres obowiazywania\b",
    r"\btermin wykonania\b",
    r"\btermin realizacji\b",
    r"\btermin zwiazania oferta\b",
    r"\btermin skladania ofert\b",
    r"\btermin otwarcia ofert\b",
    r"\bkryterium\b",
    r"\bwaga\b",
    r"\bglowny kod cpv\b",
    r"\bdodatkowy kod cpv\b",
    r"\bkod cpv\b",
    r"\bsekcja\b",
    r"\badres\b",
    r"\bemail\b",
    r"\btelefon\b",
    r"\bzamawiajacy przewiduje\b",
)


def looks_like_supplier_or_address(value: Any) -> bool:
    """Detect contractor/address rows that BZP result notices may place near lot rows."""
    text = one_line(value)
    folded = fold_text(text)
    if not folded:
        return False
    company_or_address_patterns = (
        r"\bsp\.?\s*z\s*o\.?\s*o\.?\b",
        r"\bspolka\s+z\s+ograniczona\s+odpowiedzialnoscia\b",
        r"\bsp\.?\s*j\.?\b",
        r"\bspolka\s+jawna\b",
        r"\bs\.?\s*a\.?\b",
        r"\bgmbh\b",
        r"\bltd\b",
        r"\bllc\b",
        r"\bul\.?\b",
        r"\bulica\b",
        r"\bnip\b",
        r"\bregon\b",
        r"\bkrs\b",
        r"\b\d{2}-\d{3}\b",
    )
    if any(re.search(pattern, folded, flags=re.I) for pattern in company_or_address_patterns):
        return True
    # Short capitalized names without item vocabulary are often supplier names in award sections.
    item_words = (
        "dostawa", "zakup", "pakiet", "czesc", "zadanie", "preparat", "sprzet", "aparat",
        "rekawic", "strzyk", "igla", "opatr", "odczyn", "test", "material", "wyrob",
        "lek", "nici", "szew", "cewnik", "implant", "zestaw", "roztwor", "paski", "maska",
        "waporyzator", "akcesor", "dezynfek", "chlodz", "urzadzen", "system", "odziez",
    )
    token_count = len(re.findall(r"[A-Za-zÀ-ž]+", text))
    if token_count <= 3 and not any(word in folded for word in item_words):
        if re.search(r"[A-ZĄĆĘŁŃÓŚŹŻ]{2,}|[A-ZĄĆĘŁŃÓŚŹŻ][a-ząćęłńóśźż]+", text):
            return True
    return False


def item_has_real_quantity(item: Dict[str, str]) -> bool:
    return one_line(item.get("item_quantity")) not in ("", NA) and one_line(item.get("item_uom")) not in ("", NA)


def score_item_candidate(item: Dict[str, str]) -> Tuple[int, int, int, int]:
    desc = one_line(item.get("item_desc"))
    qty = one_line(item.get("item_quantity"))
    uom = one_line(item.get("item_uom"))
    return (
        1 if item_has_real_quantity(item) else 0,
        0 if looks_like_supplier_or_address(desc) and not item_has_real_quantity(item) else 1,
        1 if desc and desc != NA else 0,
        min(len(desc), 500),
    )


def finalize_items(items: Sequence[Dict[str, str]], title: str = "", max_items: int = 0) -> List[Dict[str, str]]:
    """Clean, reject supplier rows, and keep the best row per item number.

    Award notices often repeat lot numbers once for lot descriptions and again for
    contractors. This keeps the descriptive/quantity row and drops the contractor row.
    """
    cleaned: List[Dict[str, str]] = []
    for raw_item in items:
        item = dict(raw_item)
        desc = one_line(item.get("item_desc")) or one_line(title)
        uom = one_line(item.get("item_uom"))
        qty = one_line(item.get("item_quantity"))
        if is_bad_item_candidate(desc, uom, qty):
            continue
        if looks_like_supplier_or_address(desc) and not (uom and uom != NA and qty and qty != NA):
            continue
        item["item_desc"] = clean_item_description(desc, title)
        item["item_uom"] = normalize_uom(uom) if uom and uom != NA else NA
        item["item_quantity"] = normalize_quantity(qty) if qty and qty != NA else NA
        item["item_no"] = one_line(item.get("item_no")) or str(len(cleaned) + 1)
        cleaned.append(item)

    if not cleaned:
        return []

    best_by_number: Dict[str, Dict[str, str]] = {}
    order: List[str] = []
    for item in cleaned:
        no = one_line(item.get("item_no")) or str(len(order) + 1)
        if no not in best_by_number:
            best_by_number[no] = item
            order.append(no)
            continue
        if score_item_candidate(item) > score_item_candidate(best_by_number[no]):
            best_by_number[no] = item

    result = [best_by_number[no] for no in order if no in best_by_number]
    if max_items:
        result = result[:max_items]
    return result


def is_bad_item_candidate(desc: Any, uom: Any = "", quantity: Any = "") -> bool:
    """Reject headings and contract-period snippets that look like item rows."""
    folded_desc = fold_text(desc)
    folded_uom = fold_text(uom)
    if not folded_desc:
        return True
    if folded_uom in {"dni", "dzien", "dnia", "miesiac", "miesiace", "miesiecy", "godz", "godzin", "godzina", "godziny"}:
        return True
    if looks_like_supplier_or_address(desc) and not (one_line(uom) and one_line(uom) != NA and one_line(quantity) and one_line(quantity) != NA):
        return True
    if any(re.search(pattern, folded_desc, flags=re.I) for pattern in ITEM_CONTEXT_REJECT_PATTERNS):
        return True
    if re.fullmatch(r"\d+(?:[.)])?", one_line(desc)):
        return True
    if quantity and folded_uom and folded_uom not in fold_text(UNIT_PATTERN):
        # Do not reject solely because of this heuristic; regex units are checked elsewhere.
        pass
    return False


def clean_lot_description(value: str, title: str) -> str:
    text = clean_item_description(value, title)
    text = re.sub(r"\s*4\.5\.3\.\)\s*Gł[oó]wny kod CPV:.*$", "", text, flags=re.I)
    text = re.sub(r"\s*4\.2\.6\.\)\s*Gł[oó]wny kod CPV:.*$", "", text, flags=re.I)
    text = re.sub(r"\s*Szczeg[oó]łowy opis.*?(?:SWZ|zał[aą]cznik.*)$", "", text, flags=re.I)
    text = re.sub(r"\s+", " ", text).strip(" .,:;-–")
    return text or clean_item_description(value, title)


def valid_item_or_none(item: Optional[Dict[str, str]], title: str = "") -> Optional[Dict[str, str]]:
    if not item:
        return None
    desc = one_line(item.get("item_desc")) or one_line(title)
    uom = one_line(item.get("item_uom"))
    qty = one_line(item.get("item_quantity"))
    if is_bad_item_candidate(desc, uom, qty):
        return None
    item["item_desc"] = clean_item_description(desc, title)
    if uom:
        item["item_uom"] = normalize_uom(uom)
    if qty:
        item["item_quantity"] = normalize_quantity(qty)
    return item


def clean_item_description(value: str, title: str) -> str:
    text = one_line(value)
    text = re.split(r"szczegolowy\s+opis", text, maxsplit=1, flags=re.I)[0].strip()
    text = re.sub(r"^(?:przedmiotem\s+zamowienia\s+jest|przedmiot\s+zamowienia\s+stanowi)\s+", "", fold_text(text), flags=re.I) if False else text
    text = re.sub(
        r"^(?:1[.)]?\s*)?(?:Przedmiotem\s+zam[o\u00f3]wienia\s+jest\s+|Przedmiot\s+zam[o\u00f3]wienia\s+stanowi\s+)",
        "",
        text,
        flags=re.I,
    )
    text = re.sub(r"^(?:dostawa|zakup(?:u)?|zakup\s+i\s+dostawa)\s+", "", text, flags=re.I)
    text = re.sub(r"^\d+\s*[-.)]\s*", "", text)
    text = re.sub(r"\s+", " ", text).strip(" .,:;\"'`-_/\\")
    if len(text) > 500:
        sentence_cut = re.search(r"(?<=[.!?])\s+", text[:500])
        if sentence_cut:
            text = text[: sentence_cut.end()].strip()
        else:
            text = text[:500].rsplit(" ", 1)[0].strip()
    if not text or len(text) < 3:
        text = one_line(title)
    return text


def normalize_quantity(quantity: str) -> str:
    text = one_line(quantity).replace(" ", "")
    return text.strip(".,;")


def normalize_uom(uom: str) -> str:
    return one_line(uom).strip(".,;:")


def parse_delimited_item_row(line: str) -> Optional[Dict[str, str]]:
    if "|" not in line and "\t" not in line and ";" not in line:
        return None
    parts = [one_line(part) for part in re.split(r"\t|\||;", line) if one_line(part)]
    if len(parts) < 3:
        return None
    item_no = ""
    if re.fullmatch(r"\d{1,4}[.)]?", parts[0]):
        item_no = re.sub(r"\D", "", parts.pop(0))
    qty_idx = -1
    for idx in range(len(parts) - 1, -1, -1):
        if re.fullmatch(r"\d[\d\s.,]*", parts[idx]):
            qty_idx = idx
            break
    if qty_idx < 0:
        return None
    quantity = normalize_quantity(parts[qty_idx])
    uom = ""
    desc_parts = parts[:qty_idx]
    for idx in range(qty_idx - 1, -1, -1):
        if re.fullmatch(UNIT_PATTERN, fold_text(parts[idx]), flags=re.I):
            uom = normalize_uom(parts[idx])
            desc_parts = parts[:idx]
            break
    if not uom:
        return None
    desc = clean_item_description(" ".join(desc_parts), "")
    if not desc:
        return None
    return valid_item_or_none({"item_no": item_no or "1", "item_desc": desc, "item_uom": uom, "item_quantity": quantity})


def parse_item_line(line: str, title: str) -> Optional[Dict[str, str]]:
    line = one_line(line)
    if len(line) < 8:
        return None
    folded = fold_text(line)
    if re.search(r"\b(kod\s+cpv|sekcja|termin|adres|email|telefon|kryterium|waga)\b", folded):
        return None
    if re.search(r"^(?:l\.?p\.?|lp\.?)\b", folded) or ("j.m" in folded and "ilosc" in folded):
        return None

    delimited = parse_delimited_item_row(line)
    if delimited:
        return delimited

    patterns = (
        rf"^\s*(?P<no>\d{{1,4}})[.)]?\s+(?P<desc>.+?)\s+(?P<uom>{UNIT_PATTERN})\s+(?P<qty>\d[\d\s.,]*)\s*$",
        rf"^\s*(?P<no>\d{{1,4}})[.)]?\s+(?P<desc>.+?)\s+(?P<qty>\d[\d\s.,]*)\s+(?P<uom>{UNIT_PATTERN})\s*$",
        rf"(?P<desc>.+?)\s*[-:]\s*(?P<qty>\d[\d\s.,]*)\s+(?P<uom>{UNIT_PATTERN})\s*$",
        rf"(?P<desc>.+?)\s+(?P<uom>{UNIT_PATTERN})\s*[x.]?\s*(?P<qty>\d[\d\s.,]*)\s*$",
    )
    for pattern in patterns:
        match = re.search(pattern, line, flags=re.I)
        if not match:
            continue
        desc = clean_item_description(match.group("desc"), title)
        uom = normalize_uom(match.group("uom"))
        qty = normalize_quantity(match.group("qty"))
        item_no = match.groupdict().get("no") or "1"
        if desc and uom and qty:
            return valid_item_or_none({"item_no": item_no, "item_desc": desc, "item_uom": uom, "item_quantity": qty}, title)

    folded_line = fold_text(line)
    match = re.search(r"(?P<desc>.+?)\bilosc\s+(?:opakowan|op\.?)\s*(?P<qty>\d[\d\s.,]*)\s*$", folded_line, flags=re.I)
    if match:
        original_desc = line[: max(0, match.start("desc") + len(match.group("desc")))]
        return valid_item_or_none(
            {
                "item_no": "1",
                "item_desc": clean_item_description(original_desc, title),
                "item_uom": "opak.",
                "item_quantity": normalize_quantity(match.group("qty")),
            },
            title,
        )
    return None


def extract_quantity_unit(text: str) -> Tuple[str, str]:
    folded = fold_text(text)
    if any(re.search(pattern, folded, flags=re.I) for pattern in ITEM_CONTEXT_REJECT_PATTERNS):
        return "", ""
    match = re.search(r"\bilosc\s+(?:opakowan|op\.?)\s*(?P<qty>\d[\d\s.,]*)", folded, flags=re.I)
    if match:
        desc = one_line(text[: match.start()]) if isinstance(text, str) else ""
        if not is_bad_item_candidate(desc or text, "opak.", match.group("qty")):
            return normalize_quantity(match.group("qty")), "opak."
    match = re.search(rf"(?P<qty>\d[\d\s.,]*)\s*(?P<uom>{UNIT_PATTERN})\b", folded, flags=re.I)
    if match:
        qty = normalize_quantity(match.group("qty"))
        uom = normalize_uom(match.group("uom"))
        context = folded[max(0, match.start() - 60): min(len(folded), match.end() + 60)]
        has_quantity_context = bool(
            re.search(r"\b(ilosc|ilosci|liczba|zapotrzeb|zamawian|wymagan|wielkosc|szacunkow|lacznie|razem)\b", context)
        )
        # Avoid treating product size/dosage such as `strzykawka 10 ml` as item_quantity.
        if qty and uom and has_quantity_context and not is_bad_item_candidate(text, uom, qty):
            return qty, uom
    return "", ""


def parse_compact_numbered_table(text: str, title: str) -> List[Dict[str, str]]:
    """Parse tables flattened into one line, e.g. header + `1 item szt. 10 2 item szt. 20`.

    PDF/DOCX extraction often collapses rows into a single line. We only split
    on a sequential item number (1, 2, 3, ...), so product strengths like
    `10 ml` are not mistaken for item number 10 or quantity.
    """
    compact = one_line(text)
    if not compact:
        return []
    folded = fold_text(compact)
    if not re.search(r"\b(l\.?p\.?|lp|nazwa|asortyment|j\.?m\.?|jednostka|ilosc|ilość)\b", folded, flags=re.I):
        return []

    starts = []
    expected = 1
    number_re = re.compile(r"(?<![\d,.])(?P<no>\d{1,3})(?:[.)]|\s+)(?=[^\d\s])")
    for match in number_re.finditer(compact):
        try:
            no = int(match.group("no"))
        except ValueError:
            continue
        if no == expected:
            starts.append((no, match.start(), match.end()))
            expected += 1
    if not starts:
        return []

    items: List[Dict[str, str]] = []
    seen = set()
    for idx, (no, start, body_start) in enumerate(starts):
        next_start = starts[idx + 1][1] if idx + 1 < len(starts) else len(compact)
        body = compact[body_start:next_start].strip(" .,:;|-–")
        if not body:
            continue
        item = parse_item_line(f"{no} {body}", title)
        if not item:
            continue
        key = (item.get("item_no", ""), fold_text(item.get("item_desc", "")), item.get("item_quantity", ""), fold_text(item.get("item_uom", "")))
        if key in seen:
            continue
        seen.add(key)
        items.append(item)
    return finalize_items(items, title)


def extract_table_items(text: str, title: str) -> List[Dict[str, str]]:
    raw = normalize_text(text, keep_newlines=True)
    items: List[Dict[str, str]] = []
    seen = set()
    for line in raw.splitlines():
        item = parse_item_line(line, title)
        if not item:
            continue
        key = (item.get("item_no", ""), fold_text(item.get("item_desc", "")), item.get("item_quantity", ""), fold_text(item.get("item_uom", "")))
        if key in seen:
            continue
        seen.add(key)
        items.append(item)

    if items:
        return items

    compact_items = parse_compact_numbered_table(raw, title)
    if compact_items:
        return compact_items

    compact = one_line(raw)
    candidate_pattern = re.compile(
        rf"(?P<no>\d{{1,4}})[.)]\s+(?P<body>.*?)(?=\s+\d{{1,4}}[.)]\s+|$)",
        re.I,
    )
    for row in candidate_pattern.finditer(compact):
        item = parse_item_line(f"{row.group('no')} {row.group('body')}", title)
        if item:
            key = (item.get("item_no", ""), fold_text(item.get("item_desc", "")), item.get("item_quantity", ""), fold_text(item.get("item_uom", "")))
            if key not in seen:
                seen.add(key)
                items.append(item)
    return items


def extract_named_lot_items(text: str, title: str) -> List[Dict[str, str]]:
    """Extract package/lot names from Część/Pakiet/Zadanie sections.

    BZP result notices often contain only package-level names, while the full
    quantities live in attachments. This at least prevents the scraper from
    falling back to random period/date lines and gives one row per package.
    """
    lines = normalize_text(text, keep_newlines=True).splitlines()
    items: List[Dict[str, str]] = []
    seen = set()
    current_no = ""
    for index, line in enumerate(lines):
        clean = line.strip().lstrip("#").strip()
        folded = fold_text(clean)
        lot_match = re.search(r"\b(?:czesc|pakiet|zadanie)\s*(?:nr|numer)?\s*[:#-]?\s*(?P<no>\d{1,4})\b", folded, flags=re.I)
        if lot_match:
            current_no = lot_match.group("no")

        name = ""
        if "nazwa" in folded and ":" in clean and current_no:
            name = clean.split(":", 1)[1].strip()
        elif re.search(r"(?:krotki opis przedmiotu zamowienia|opis przedmiotu zamowienia)", folded) and current_no:
            for next_line in lines[index + 1 : index + 8]:
                if not next_line.strip():
                    continue
                if is_numbered_heading(next_line):
                    break
                name = next_line.strip().lstrip("#").strip()
                break
        elif current_no and re.match(r"^\s*(?:pakiet|zadanie)\s*(?:nr|numer)?\s*\d+\b", clean, flags=re.I):
            name = clean

        if not name:
            continue
        desc = clean_lot_description(name, title)
        if is_bad_item_candidate(desc):
            continue
        key = (current_no or str(len(items) + 1), fold_text(desc))
        if key in seen:
            continue
        seen.add(key)
        items.append({"item_no": current_no or str(len(items) + 1), "item_desc": desc, "item_uom": NA, "item_quantity": NA})

    if items:
        return items

    # Some snippets collapse all lots into one line. Parse compact Część/Pakiet/Zadanie blocks too.
    compact = one_line(text)
    compact_lot_re = re.compile(
        r"\b(?:czesc|pakiet|zadanie)\s*(?:nr|numer)?\s*[:#-]?\s*(?P<no>\d{1,4})\b(?P<body>.*?)(?=\b(?:czesc|pakiet|zadanie)\s*(?:nr|numer)?\s*[:#-]?\s*\d{1,4}\b|$)",
        re.I,
    )
    for match in compact_lot_re.finditer(fold_text(compact)):
        no = match.group("no")
        # Use the original compact string slice when possible so Polish characters are retained.
        body_start = match.start("body")
        body_end = match.end("body")
        original_body = compact[body_start:body_end] if body_end <= len(compact) else match.group("body")
        if len(original_body.strip(" -:;,.")) < 3:
            continue
        desc = clean_lot_description(original_body, title)
        if desc == one_line(title) or is_bad_item_candidate(desc):
            continue
        key = (no, fold_text(desc))
        if key in seen:
            continue
        seen.add(key)
        items.append({"item_no": no, "item_desc": desc, "item_uom": NA, "item_quantity": NA})

    return finalize_items(items, title)


def extract_lot_items(html_text: str, title: str) -> List[Dict[str, str]]:
    sections = extract_numbered_sections(html_text, ("4.2.2", "4.5.1"))
    if len(sections) <= 1:
        return []
    items: List[Dict[str, str]] = []
    for idx, section in enumerate(sections, start=1):
        table_items = extract_table_items(section, title)
        if table_items:
            for item in table_items:
                item["item_no"] = item.get("item_no") or str(idx)
                items.append(item)
            continue
        desc = clean_lot_description(section, title)
        qty, uom = extract_quantity_unit(section)
        if is_bad_item_candidate(desc, uom, qty):
            continue
        items.append(
            {
                "item_no": str(idx),
                "item_desc": desc,
                "item_uom": uom or NA,
                "item_quantity": qty or NA,
            }
        )
    return finalize_items(items, title)


def extract_items(html_text: str, description: str, title: str, max_items: int = 0) -> List[Dict[str, str]]:
    sources = [description, html_text]
    items: List[Dict[str, str]] = []
    seen = set()
    for source in sources:
        for item in extract_table_items(source, title):
            key = (item.get("item_no", ""), fold_text(item.get("item_desc", "")), item.get("item_quantity", ""), fold_text(item.get("item_uom", "")))
            if key in seen:
                continue
            seen.add(key)
            items.append(item)
            if max_items and len(items) >= max_items:
                return items
    if items:
        return finalize_items(items, title, max_items=max_items)

    named_lot_items = extract_named_lot_items(html_text, title)
    if named_lot_items:
        return finalize_items(named_lot_items, title, max_items=max_items)

    lot_items = extract_lot_items(html_text, title)
    if lot_items:
        return finalize_items(lot_items, title, max_items=max_items)

    qty, uom = extract_quantity_unit(description)
    desc = clean_lot_description(description or title, title)
    fallback = {
        "item_no": "1",
        "item_desc": desc,
        "item_uom": uom or NA,
        "item_quantity": qty or NA,
    }
    if is_bad_item_candidate(fallback["item_desc"], fallback["item_uom"], fallback["item_quantity"]):
        fallback = {"item_no": "1", "item_desc": clean_item_description(title or description, title), "item_uom": NA, "item_quantity": NA}
    return finalize_items([fallback], title, max_items=max_items) or [fallback]


def item_detail_score(items: Sequence[Dict[str, str]]) -> Tuple[int, int, int, int]:
    good_items = [item for item in items if not looks_like_supplier_or_address(item.get("item_desc"))]
    with_qty_uom = sum(1 for item in good_items if item_has_real_quantity(item))
    with_desc = sum(1 for item in good_items if one_line(item.get("item_desc")) not in ("", NA))
    total_desc_len = sum(len(one_line(item.get("item_desc"))) for item in good_items)
    return with_qty_uom, len(good_items), with_desc, total_desc_len


def choose_better_items(primary: List[Dict[str, str]], alternative: List[Dict[str, str]]) -> List[Dict[str, str]]:
    primary_clean = finalize_items(primary) or primary
    alternative_clean = finalize_items(alternative) or alternative
    if not alternative_clean:
        return primary_clean
    if not primary_clean:
        return alternative_clean
    return alternative_clean if item_detail_score(alternative_clean) > item_detail_score(primary_clean) else primary_clean


def build_procedure_urls(tender_id: str) -> List[str]:
    tender_id = one_line(tender_id)
    return [template.format(tender_id=tender_id) for template in PROCEDURE_URLS if tender_id]


def is_interesting_document_name(name_or_url: str) -> bool:
    folded = fold_text(unquote(name_or_url))
    if not folded:
        return False
    if not re.search(r"\.(xlsx?|xlsm|xltx|xltm|csv|docx|pdf|zip)(?:$|[?#])", folded):
        return False
    return any(re.search(pattern, folded, flags=re.I) for pattern in ITEM_DOCUMENT_NAME_PATTERNS)


def extract_document_links_from_html(page_url: str, html_text: str) -> List[Tuple[str, str]]:
    """Extract likely public document links from normal HTML, Angular markup, and embedded JSON.

    e-Zamówienia procedure pages are not always simple <a href="file.xlsx"> links.
    Depending on how the page is rendered, files can appear as anchors whose URL is
    a download endpoint, data-url attributes, or JSON fragments containing file names.
    This extractor intentionally keeps any URL/relative link when the visible label
    or nearby context looks like SWZ/OPZ/formularz/asortyment/cenowy material.
    """
    links: List[Tuple[str, str]] = []
    seen = set()

    def add_link(href: str, name: str = "") -> None:
        href = html.unescape(one_line(href)).strip()
        name = html.unescape(one_line(name)).strip()
        if not href or href.startswith(("javascript:", "mailto:", "tel:")):
            return
        # Skip application shell routes unless they directly identify a file/download.
        folded_href = fold_text(href)
        folded_name = fold_text(name)
        combined = " ".join([folded_name, folded_href])
        has_file_extension = bool(re.search(r"\.(xlsx?|xlsm|xltx|xltm|csv|docx|pdf|zip)(?:$|[?#])", combined, flags=re.I))
        has_download_marker = bool(re.search(r"\b(download|pobierz|file|document|attachment|zalacznik|załącznik|media)\b", combined, flags=re.I))
        has_interesting_name = any(re.search(pattern, combined, flags=re.I) for pattern in ITEM_DOCUMENT_NAME_PATTERNS)
        if not (has_file_extension or (has_download_marker and has_interesting_name)):
            return
        url = urljoin(page_url, href)
        if url not in seen:
            seen.add(url)
            links.append((url, name or unquote(Path(urlparse(url).path).name) or url))

    # 1) Standard anchors and data-url style attributes.
    for match in re.finditer(
        r"<(?:a|button|div|span)[^>]+(?:href|data-href|data-url|data-download-url|ng-href)=[\"'](?P<href>[^\"']+)[\"'][^>]*>(?P<label>.*?)</(?:a|button|div|span)>",
        html_text,
        flags=re.I | re.S,
    ):
        label = normalize_text(match.group("label"))
        add_link(match.group("href"), label)

    # 2) Any href-like attribute. Use surrounding context as the label.
    for match in re.finditer(r"(?:href|data-href|data-url|data-download-url|ng-href)=[\"'](?P<href>[^\"']+)[\"']", html_text, flags=re.I):
        start = max(0, match.start() - 300)
        end = min(len(html_text), match.end() + 300)
        context = normalize_text(html_text[start:end])
        add_link(match.group("href"), context)

    # 3) Absolute URLs in text or JSON.
    for match in re.finditer(r"https?://[^\s\"'<>\\)]+", html_text, flags=re.I):
        start = max(0, match.start() - 220)
        end = min(len(html_text), match.end() + 220)
        context = normalize_text(html_text[start:end])
        add_link(match.group(0), context)

    # 4) JSON-style fileName/url pairs. This catches many SPA payloads if embedded.
    json_pair_patterns = (
        r"[\"'](?P<name_key>fileName|filename|name|originalFileName|documentName|nazwaPliku)[\"']\s*:\s*[\"'](?P<name>[^\"']+)[\"'](?P<mid>.{0,800}?)[\"'](?P<url_key>url|href|downloadUrl|fileUrl|link)[\"']\s*:\s*[\"'](?P<href>[^\"']+)[\"']",
        r"[\"'](?P<url_key>url|href|downloadUrl|fileUrl|link)[\"']\s*:\s*[\"'](?P<href>[^\"']+)[\"'](?P<mid>.{0,800}?)[\"'](?P<name_key>fileName|filename|name|originalFileName|documentName|nazwaPliku)[\"']\s*:\s*[\"'](?P<name>[^\"']+)[\"']",
    )
    for pattern in json_pair_patterns:
        for match in re.finditer(pattern, html_text, flags=re.I | re.S):
            add_link(match.group("href"), match.group("name"))

    return links


def extract_page_urls_from_notice_text(*texts: str) -> List[str]:
    """Find candidate pages from notice/original-notice text that may host attachments."""
    urls: List[str] = []
    seen = set()
    for text in texts:
        for match in re.finditer(r"https?://[^\s\"'<>\\)]+", text or "", flags=re.I):
            url = html.unescape(match.group(0)).rstrip(".,;)]")
            folded = fold_text(url)
            if not url or url in seen:
                continue
            # Notice/detail pages and hosted procurement/BIP pages are useful; static image/css/js URLs are not.
            if re.search(r"\.(?:png|jpe?g|gif|svg|css|js)(?:$|[?#])", folded):
                continue
            if any(marker in folded for marker in ("ezamowienia.gov.pl", "bip", "zamow", "przetarg", "platformazakupowa", "ezamawiajacy")):
                seen.add(url)
                urls.append(url)
    return urls


def fetch_document_links_from_pages(session: requests.Session, page_urls: Sequence[str]) -> List[Tuple[str, str]]:
    links: List[Tuple[str, str]] = []
    seen = set()
    for page_url in page_urls:
        try:
            response = session.get(page_url, timeout=45)
            response.raise_for_status()
        except Exception as exc:
            logging.debug("Could not fetch procedure/source page %s: %s", page_url, exc)
            continue
        found = extract_document_links_from_html(page_url, response.text)
        logging.debug("Document-link scan on %s found %s candidate links", page_url, len(found))
        for url, name in found:
            if url in seen:
                continue
            seen.add(url)
            links.append((url, name))
    return links


def fetch_procedure_document_links(session: requests.Session, tender_id: str, extra_page_urls: Optional[Sequence[str]] = None) -> List[Tuple[str, str]]:
    pages = build_procedure_urls(tender_id)
    if extra_page_urls:
        for url in extra_page_urls:
            if url not in pages:
                pages.append(url)
    return fetch_document_links_from_pages(session, pages)


def fetch_procedure_document_links_with_playwright(tender_id: str, wait_ms: int = 5000) -> List[Tuple[str, str]]:
    """Render public e-Zamówienia procedure pages and collect document URLs.

    Some procedure pages are JavaScript applications. Plain requests can see
    only the app shell, while the real SWZ/OPZ/formularz documents are loaded
    by the browser. This optional path executes the page and collects anchors,
    resource URLs, and document-like routes.
    """
    try:
        from playwright.sync_api import sync_playwright  # type: ignore
    except Exception as exc:
        logging.warning(
            "Playwright document rendering was requested but Playwright is not installed. "
            "Install it with: python -m pip install playwright && python -m playwright install chromium. Error: %s",
            exc,
        )
        return []

    collected: List[Tuple[str, str]] = []
    seen = set()
    try:
        with sync_playwright() as playwright:
            browser = playwright.chromium.launch(headless=True)
            context = browser.new_context(accept_downloads=True, locale="pl-PL")
            page = context.new_page()
            for page_url in build_procedure_urls(tender_id):
                try:
                    page.goto(page_url, wait_until="networkidle", timeout=70000)
                    page.wait_for_timeout(wait_ms)
                    for _ in range(5):
                        page.mouse.wheel(0, 3000)
                        page.wait_for_timeout(400)
                except Exception as exc:
                    logging.debug("Playwright could not render %s: %s", page_url, exc)
                    continue

                candidates: List[Tuple[str, str]] = []
                try:
                    candidates.extend(extract_document_links_from_html(page_url, page.content()))
                except Exception:
                    pass

                try:
                    resource_urls = page.evaluate("() => performance.getEntriesByType('resource').map(e => e.name)")
                except Exception:
                    resource_urls = []
                for resource_url in resource_urls or []:
                    if not isinstance(resource_url, str):
                        continue
                    name = unquote(Path(urlparse(resource_url).path).name) or resource_url
                    combined = " ".join([name, resource_url])
                    if is_interesting_document_name(combined) or re.search(r"\b(document|attachment|file|download|tenderdocument)\b", fold_text(combined)):
                        candidates.append((resource_url, name))

                try:
                    dom_items = page.evaluate(
                        """() => Array.from(document.querySelectorAll('a,button,[role=button]')).map((el, idx) => ({
                            idx,
                            text: (el.innerText || el.textContent || '').trim(),
                            href: el.href || el.getAttribute('href') || '',
                            title: el.getAttribute('title') || '',
                            aria: el.getAttribute('aria-label') || ''
                        }))"""
                    )
                except Exception:
                    dom_items = []
                for item in dom_items or []:
                    href = str(item.get("href") or "")
                    label = one_line(" ".join([str(item.get("text") or ""), str(item.get("title") or ""), str(item.get("aria") or "")]))
                    if href:
                        url = urljoin(page_url, href)
                        combined = " ".join([label, url])
                        if is_interesting_document_name(combined) or re.search(r"\b(pobierz|download|tenderdocument|document|zalacznik|załącznik)\b", fold_text(combined)):
                            candidates.append((url, label or unquote(Path(urlparse(url).path).name) or url))

                for url, name in candidates:
                    if not url or url in seen:
                        continue
                    seen.add(url)
                    collected.append((url, name))
            try:
                context.close()
                browser.close()
            except Exception:
                pass
    except Exception as exc:
        logging.warning("Playwright document rendering failed for tender %s: %s", tender_id, exc)
    if collected:
        logging.info("Playwright found %s candidate document links for tender %s", len(collected), tender_id)
    return collected


def safe_document_filename(url: str, name: str) -> str:
    url_name = unquote(Path(urlparse(url).path).name)
    label_name = one_line(name)
    candidate = url_name or label_name or "document"
    # If the download endpoint has no extension but the visible label does, keep the label extension.
    label_ext_match = re.search(r"\.(xlsx?|xlsm|xltx|xltm|csv|docx|pdf|zip)\b", label_name, flags=re.I)
    if label_ext_match and not re.search(r"\.(xlsx?|xlsm|xltx|xltm|csv|docx|pdf|zip)\b", candidate, flags=re.I):
        candidate = f"{candidate}.{label_ext_match.group(1).lower()}"
    candidate = re.sub(r"[^A-Za-z0-9._ -]+", "_", candidate).strip(" ._")
    return candidate[:180] or "document"


def download_document(session: requests.Session, url: str, name: str, cache_dir: Path) -> Optional[Path]:
    cache_dir.mkdir(parents=True, exist_ok=True)
    filename = safe_document_filename(url, name)
    path = cache_dir / filename
    if path.exists() and path.stat().st_size > 0:
        return path
    try:
        response = session.get(url, timeout=90)
        response.raise_for_status()
        content = response.content
    except Exception as exc:
        logging.debug("Could not download document %s: %s", url, exc)
        return None
    if not content:
        return None
    path.write_bytes(content)
    return path


def extract_text_from_docx(path: Path) -> str:
    try:
        with zipfile.ZipFile(path) as zf:
            xml_parts = []
            for name in zf.namelist():
                if name.startswith("word/") and name.endswith(".xml") and ("document" in name or "table" in name):
                    xml_parts.append(zf.read(name).decode("utf-8", errors="ignore"))
    except Exception:
        return ""
    texts: List[str] = []
    for xml_text in xml_parts:
        # Preserve table-cell and paragraph boundaries enough for line/table parsing.
        xml_text = re.sub(r"</w:tc>", "\t", xml_text)
        xml_text = re.sub(r"</w:tr>", "\n", xml_text)
        xml_text = re.sub(r"</w:p>", "\n", xml_text)
        xml_text = re.sub(r"<w:tab\s*/>", "\t", xml_text)
        xml_text = re.sub(r"<w:br\s*/>", "\n", xml_text)
        xml_text = re.sub(r"<[^>]+>", "", xml_text)
        texts.append(html.unescape(xml_text))
    return normalize_text("\n".join(texts), keep_newlines=True)


def extract_text_from_xlsx(path: Path) -> str:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception:
        return ""
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return ""
    lines: List[str] = []
    try:
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                values = [one_line(cell) for cell in row if one_line(cell)]
                if values:
                    lines.append("\t".join(values))
    finally:
        try:
            workbook.close()
        except Exception:
            pass
    return "\n".join(lines)


def extract_text_from_xls(path: Path) -> str:
    """Read legacy .xls files if xlrd is installed."""
    try:
        import xlrd  # type: ignore
    except Exception:
        logging.debug("xlrd is not installed; cannot parse legacy .xls file %s", path)
        return ""
    try:
        workbook = xlrd.open_workbook(str(path), on_demand=True)
    except Exception:
        return ""
    lines: List[str] = []
    try:
        for sheet in workbook.sheets():
            for row_idx in range(sheet.nrows):
                values = [one_line(sheet.cell_value(row_idx, col_idx)) for col_idx in range(sheet.ncols)]
                values = [value for value in values if value]
                if values:
                    lines.append("\t".join(values))
    finally:
        try:
            workbook.release_resources()
        except Exception:
            pass
    return "\n".join(lines)


def extract_text_from_csv_file(path: Path) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1250", "latin-1"):
        try:
            return path.read_text(encoding=encoding, errors="ignore")
        except Exception:
            continue
    return ""


def extract_text_from_pdf(path: Path) -> str:
    """Extract text from text-based PDFs without OCR.

    OCR is intentionally not attempted. If a PDF is scanned/image-only, this
    returns blank so the scraper can continue safely.
    """
    # Try pypdf, pdfminer.six, then PyPDF2. These are optional dependencies.
    try:
        from pypdf import PdfReader  # type: ignore

        reader = PdfReader(str(path))
        pages = []
        for page in reader.pages[:80]:
            pages.append(page.extract_text() or "")
        text = "\n".join(pages)
        if one_line(text):
            return normalize_text(text, keep_newlines=True)
    except Exception:
        pass
    try:
        from pdfminer.high_level import extract_text as pdfminer_extract_text  # type: ignore

        text = pdfminer_extract_text(str(path), maxpages=80)
        if one_line(text):
            return normalize_text(text, keep_newlines=True)
    except Exception:
        pass
    try:
        from PyPDF2 import PdfReader as PyPDF2Reader  # type: ignore

        reader = PyPDF2Reader(str(path))
        pages = []
        for page in reader.pages[:80]:
            pages.append(page.extract_text() or "")
        text = "\n".join(pages)
        if one_line(text):
            return normalize_text(text, keep_newlines=True)
    except Exception:
        pass
    return ""


def extract_text_from_html_file(path: Path) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1250", "latin-1"):
        try:
            return normalize_text(path.read_text(encoding=encoding, errors="ignore"), keep_newlines=True)
        except Exception:
            continue
    return ""


def extract_texts_from_zip(path: Path, max_members: int = 80) -> List[str]:
    texts: List[str] = []
    try:
        with zipfile.ZipFile(path) as zf:
            members = [m for m in zf.namelist() if not m.endswith("/")]
            # Prefer assortment/price/form attachments, not signatures or huge legal boilerplate.
            members.sort(key=lambda name: (0 if any(re.search(p, fold_text(name), flags=re.I) for p in ITEM_DOCUMENT_NAME_PATTERNS) else 1, len(name)))
            for member in members[:max_members]:
                suffix = Path(member).suffix.lower()
                if suffix not in {".docx", ".xlsx", ".xlsm", ".xltx", ".xltm", ".csv", ".pdf", ".html", ".htm", ".xls"}:
                    continue
                try:
                    data = zf.read(member)
                except Exception:
                    continue
                temp_name = re.sub(r"[^A-Za-z0-9._-]+", "_", Path(member).name) or f"member{len(texts)}{suffix}"
                temp_path = path.parent / f"__zip_{path.stem}_{temp_name}"
                try:
                    temp_path.write_bytes(data)
                    text = extract_text_from_document(temp_path)
                    if one_line(text):
                        texts.append(text)
                finally:
                    try:
                        temp_path.unlink()
                    except Exception:
                        pass
    except Exception as exc:
        logging.debug("Could not read ZIP document %s: %s", path, exc)
    return texts


def extract_text_from_document(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == ".docx":
        return extract_text_from_docx(path)
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return extract_text_from_xlsx(path)
    if suffix == ".xls":
        return extract_text_from_xls(path)
    if suffix == ".csv":
        return extract_text_from_csv_file(path)
    if suffix == ".pdf":
        return extract_text_from_pdf(path)
    if suffix in {".html", ".htm"}:
        return extract_text_from_html_file(path)
    if suffix == ".zip":
        return "\n".join(extract_texts_from_zip(path))
    return ""


def extract_items_from_candidate_pages(
    session: requests.Session,
    page_urls: Sequence[str],
    title: str,
    max_items: int = 0,
) -> List[Dict[str, str]]:
    """Try to parse visible HTML pages themselves as a last-mile item source."""
    items: List[Dict[str, str]] = []
    seen = set()
    for page_url in page_urls:
        try:
            response = session.get(page_url, timeout=45)
            response.raise_for_status()
        except Exception:
            continue
        text = normalize_text(response.text, keep_newlines=True)
        page_items = extract_table_items(text, title)
        if not page_items:
            page_items = extract_named_lot_items(text, title)
        for item in page_items:
            key = (item.get("item_no", ""), fold_text(item.get("item_desc", "")), item.get("item_quantity", ""), fold_text(item.get("item_uom", "")))
            if key in seen:
                continue
            seen.add(key)
            items.append(item)
            if max_items and len(items) >= max_items:
                return items
    return items


def extract_items_from_procedure_documents(
    session: requests.Session,
    tender_id: str,
    title: str,
    cache_dir: Path,
    max_documents: int = 5,
    max_items: int = 0,
    extra_page_urls: Optional[Sequence[str]] = None,
    use_playwright: bool = False,
) -> List[Dict[str, str]]:
    source_pages = build_procedure_urls(tender_id)
    if extra_page_urls:
        for page_url in extra_page_urls:
            if page_url not in source_pages:
                source_pages.append(page_url)

    page_items = extract_items_from_candidate_pages(session, source_pages, title, max_items=max_items)
    if item_detail_score(page_items)[0] > 0:
        logging.info("Procedure/source page item extraction found %s item rows for tender %s", len(page_items), tender_id)
        return page_items

    links = fetch_procedure_document_links(session, tender_id, extra_page_urls=extra_page_urls)
    if use_playwright:
        rendered_links = fetch_procedure_document_links_with_playwright(tender_id)
        existing_urls = {url for url, _name in links}
        for url, name in rendered_links:
            if url not in existing_urls:
                links.append((url, name))
                existing_urls.add(url)
    if not links:
        logging.info("No public document links found for tender %s", tender_id)
        return page_items
    logging.info("Found %s public document links for tender %s; inspecting up to %s", len(links), tender_id, max_documents)
    items: List[Dict[str, str]] = []
    seen = set()
    tender_cache_dir = cache_dir / re.sub(r"[^A-Za-z0-9._-]+", "_", tender_id or "no_tender_id")
    inspected = 0
    for url, name in links:
        if inspected >= max(1, max_documents):
            break
        path = download_document(session, url, name, tender_cache_dir)
        if not path:
            continue
        inspected += 1
        text = extract_text_from_document(path)
        if not text:
            logging.debug("Downloaded %s but no readable text was extracted", path)
            continue
        doc_items = extract_table_items(text, title)
        if not doc_items:
            # Some spreadsheets/documents have package rows with item names but no simple qty/uom pattern.
            doc_items = extract_named_lot_items(text, title)
        logging.debug("Document %s produced %s item candidates", path.name, len(doc_items))
        for item in doc_items:
            key = (item.get("item_no", ""), fold_text(item.get("item_desc", "")), item.get("item_quantity", ""), fold_text(item.get("item_uom", "")))
            if key in seen:
                continue
            seen.add(key)
            items.append(item)
            if max_items and len(items) >= max_items:
                logging.info("Document item extraction found %s item rows for tender %s", len(items), tender_id)
                return items
    if items:
        logging.info("Document item extraction found %s item rows for tender %s", len(items), tender_id)
    else:
        logging.info("Documents were found for tender %s, but no structured item rows were readable", tender_id)
    return items


def item_rows_have_quantity_and_uom(rows: Sequence[Dict[str, str]]) -> bool:
    return any(
        one_line(row.get("item_quantity")) not in ("", NA)
        and one_line(row.get("item_uom")) not in ("", NA)
        for row in rows
    )


def mapped_rows_need_document_items(rows: Sequence[Dict[str, str]]) -> bool:
    if not rows:
        return True
    return not item_rows_have_quantity_and_uom(rows)


def build_notice_url(notice: Dict[str, Any]) -> str:
    object_id = one_line(notice.get("objectId"))
    if object_id:
        return NOTICE_DETAIL_URL.format(object_id=object_id)
    return ""


def calc_item_unit_price(item_quantity: str, item_award: str, awarded_value: str) -> str:
    quantity = parse_decimal(item_quantity)
    total = parse_decimal(item_award) or parse_decimal(awarded_value)
    if quantity and total and quantity > 0:
        return format_decimal(total / quantity)
    return ""


def merge_original_notice(notice: Dict[str, Any], original_notice: Optional[Dict[str, Any]]) -> Tuple[str, str]:
    html_text = get_html_text(notice)
    if original_notice and original_notice is not notice:
        original_html = get_html_text(original_notice)
        if len(original_html) > len(html_text):
            description = extract_description(original_html, one_line(original_notice.get("orderObject")))
            return original_html, description
    return html_text, extract_description(html_text, one_line(notice.get("orderObject")))


def map_notice_to_rows(
    notice: Dict[str, Any],
    query_text: str,
    original_notice: Optional[Dict[str, Any]] = None,
    max_items_per_notice: int = 0,
    contract_period_inclusive: bool = False,
    document_items: Optional[List[Dict[str, str]]] = None,
) -> List[Dict[str, str]]:
    notice_type = one_line(notice.get("noticeType")) or one_line(notice.get("type")) or "Unknown"
    html_text, description = merge_original_notice(notice, original_notice)
    own_html_text = get_html_text(notice)
    title = one_line(notice.get("orderObject")) or first_section(html_text, ("2.3",))
    cpv = one_line(notice.get("cpvCode")) or extract_cpv_from_text(html_text)
    if original_notice and not cpv:
        cpv = one_line(original_notice.get("cpvCode")) or extract_cpv_from_text(get_html_text(original_notice))

    currency, amount, awarded_currency, awarded_value, item_award = extract_money(notice_type, own_html_text)
    supplier = extract_supplier(notice, own_html_text)
    awarded_date = extract_awarded_date(notice, own_html_text, notice_type)
    contract_period = extract_contract_period(
        own_html_text,
        notice_type,
        awarded_date=awarded_date,
        inclusive=contract_period_inclusive,
    )
    items = extract_items(html_text, description, title, max_items=max_items_per_notice)
    if document_items:
        items = choose_better_items(items, document_items)
        if max_items_per_notice:
            items = items[:max_items_per_notice]

    rows: List[Dict[str, str]] = []
    notice_number = one_line(notice.get("noticeNumber"))
    tender_id = one_line(notice.get("tenderId"))
    for index, item in enumerate(items, start=1):
        item_no = one_line(item.get("item_no")) or str(index)
        item_desc = one_line(item.get("item_desc")) or title or description
        item_uom = one_line(item.get("item_uom"))
        item_quantity = one_line(item.get("item_quantity"))
        item_unit_price = calc_item_unit_price(item_quantity, item_award, awarded_value)
        row = {
            "source": "eZamowienia BZP API",
            "country": "Poland",
            "country_cc": "PL",
            "publication_date": date_only(notice.get("publicationDate")),
            "closing_date": extract_closing_date(notice, html_text),
            "title": title,
            "description": description or title,
            "buyer": one_line(notice.get("organizationName")),
            "classification": cpv,
            "status": status_for_notice_type(notice_type),
            "currency": currency,
            "amount": amount,
            "awarding_supplier_name": supplier,
            "awarded_currency": awarded_currency,
            "awarded_value": awarded_value,
            "awarded_date": awarded_date,
            "contract_period": contract_period,
            "item_no": item_no,
            "item_desc": item_desc,
            "item_uom": item_uom,
            "item_quantity": item_quantity,
            "item_unit_price": item_unit_price,
            "item_award": item_award,
            "notice_id": notice_number or tender_id,
            "notice_url": build_notice_url(notice),
            "tender_id": tender_id,
            "notice_type": notice_type,
            "query_text": query_text,
            "scraped_at": utc_now_iso(),
            "dedup_key": "|".join(
                part
                for part in ["poland_ezamowienia", tender_id, notice_number, item_no, fold_text(item_desc)[:90]]
                if part
            ),
        }
        rows.append(fill_no_blanks(row))
    return rows


def fill_no_blanks(row: Dict[str, Any]) -> Dict[str, str]:
    return {field: nonblank(row.get(field)) for field in BASE_FIELDS}


def fill_no_blanks_all(row: Dict[str, Any]) -> Dict[str, str]:
    return {field: nonblank(row.get(field)) for field in FIELDNAMES}


def is_probably_non_linguistic(text: str) -> bool:
    if text == NA:
        return True
    if re.fullmatch(r"[\d\s.,:/+\-TZ]+", text):
        return True
    if text.startswith("http://") or text.startswith("https://"):
        return True
    if re.fullmatch(r"[A-Z]{2,4}", text):
        return True
    return False


class TranslationCache:
    def __init__(self, enabled: bool = True, sleep_seconds: float = 0.15, source: str = "pl", target: str = "en") -> None:
        self.enabled = enabled
        self.sleep_seconds = sleep_seconds
        self.cache: Dict[str, str] = {}
        self.translator = None
        if enabled:
            try:
                from deep_translator import GoogleTranslator  # type: ignore
            except ImportError as exc:
                raise SystemExit(
                    "deep-translator is required for English columns. Install it with: "
                    "python -m pip install deep-translator\n"
                    "Or rerun with --no-translate."
                ) from exc
            self.translator = GoogleTranslator(source=source, target=target)

    def split_for_translation(self, text: str, limit: int = 4400) -> List[str]:
        if len(text) <= limit:
            return [text]
        chunks: List[str] = []
        remaining = text
        while len(remaining) > limit:
            cut = max(remaining.rfind(". ", 0, limit), remaining.rfind("; ", 0, limit), remaining.rfind(" ", 0, limit))
            if cut < 1000:
                cut = limit
            chunks.append(remaining[:cut].strip())
            remaining = remaining[cut:].strip()
        if remaining:
            chunks.append(remaining)
        return chunks

    def translate(self, value: Any) -> str:
        text = nonblank(value)
        if not self.enabled or is_probably_non_linguistic(text):
            return text
        cached = self.cache.get(text)
        if cached is not None:
            return cached
        try:
            assert self.translator is not None
            translated_parts = []
            for chunk in self.split_for_translation(text):
                translated = self.translator.translate(chunk)
                translated_parts.append(one_line(translated) or chunk)
                if self.sleep_seconds:
                    time.sleep(self.sleep_seconds)
            result = one_line(" ".join(translated_parts)) or text
        except Exception as exc:
            logging.warning("Translation failed; keeping original text. Error: %s", exc)
            result = text
        self.cache[text] = result
        return result


def add_translation_columns(row: Dict[str, str], translator: TranslationCache, fields: Sequence[str]) -> Dict[str, str]:
    for field in TRANSLATE_SOURCE_FIELDS:
        target = f"{field}_en"
        if field in fields:
            row[target] = translator.translate(row.get(field, NA))
        else:
            row[target] = nonblank(row.get(field, NA))
    return fill_no_blanks_all(row)


def read_all_notices(
    session: requests.Session,
    notice_types: Sequence[str],
    date_from: str,
    date_to: str,
    page_size: int,
    window_days: int,
    request_sleep: float,
) -> List[Tuple[Dict[str, Any], str]]:
    collected: List[Tuple[Dict[str, Any], str]] = []
    for notice_type in notice_types:
        for window_from, window_to, page, notices in iter_notices(
            session,
            notice_type=notice_type,
            date_from=date_from,
            date_to=date_to,
            page_size=page_size,
            window_days=window_days,
            request_sleep=request_sleep,
        ):
            query_text = f"live:{notice_type}:{window_from}..{window_to}:page={page}"
            logging.info("Fetched %s notices for %s", len(notices), query_text)
            for notice in notices:
                collected.append((notice, query_text))
    return collected


def write_csv(path: Path, rows: Iterable[Dict[str, str]]) -> int:
    path.parent.mkdir(parents=True, exist_ok=True)
    count = 0
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=FIELDNAMES, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(fill_no_blanks_all(row))
            count += 1
    return count


def default_original_cache_path(output: Path) -> Path:
    return output.with_suffix(output.suffix + ".original_contracts.json")


def original_contract_from_date(date_from: str, lookback_days: int) -> str:
    return (parse_iso_date(date_from) - timedelta(days=max(0, int(lookback_days)))).isoformat()


def load_original_contract_cache(path: Path, expected_from: str, expected_to: str) -> Optional[Dict[str, Dict[str, Any]]]:
    if not path.exists() or path.stat().st_size == 0:
        return None
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception as exc:
        logging.warning("Could not read original ContractNotice cache %s: %s", path, exc)
        return None
    if not isinstance(payload, dict):
        return None
    meta = payload.get("meta") if isinstance(payload.get("meta"), dict) else {}
    if meta.get("date_from") != expected_from or meta.get("date_to") != expected_to:
        logging.info(
            "Original ContractNotice cache date range does not match this run; rebuilding. Cache=%s..%s, expected=%s..%s",
            meta.get("date_from"),
            meta.get("date_to"),
            expected_from,
            expected_to,
        )
        return None
    notices = payload.get("notices")
    if not isinstance(notices, dict):
        return None
    return {str(k): v for k, v in notices.items() if isinstance(v, dict)}


def save_original_contract_cache(path: Path, notices_by_tender_id: Dict[str, Dict[str, Any]], date_from: str, date_to: str) -> None:
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        temp_path = path.with_suffix(path.suffix + ".tmp")
        payload = {
            "meta": {
                "date_from": date_from,
                "date_to": date_to,
                "notice_type": "ContractNotice",
                "medical_only": True,
                "count": len(notices_by_tender_id),
                "created_at": utc_now_iso(),
            },
            "notices": notices_by_tender_id,
        }
        temp_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
        temp_path.replace(path)
    except Exception as exc:
        logging.warning("Could not save original ContractNotice cache %s: %s", path, exc)


def collect_original_contract_notice_map(
    session: requests.Session,
    date_from: str,
    date_to: str,
    page_size: int,
    window_days: int,
    request_sleep: float,
    cache_path: Path,
    refresh_cache: bool = False,
) -> Dict[str, Dict[str, Any]]:
    """Fetch medical ContractNotice records so award notices can use better item/lot text.

    TenderResultNotice and ContractPerformingNotice often only say "Pakiet nr X" or
    "details in SWZ attachment". The linked/original ContractNotice usually has the
    more useful 4.2.2 lot descriptions, so we build a tender_id -> ContractNotice map.
    """
    if not refresh_cache:
        cached = load_original_contract_cache(cache_path, date_from, date_to)
        if cached is not None:
            logging.info("Loaded %s original ContractNotice records from cache: %s", len(cached), cache_path)
            return cached

    logging.info(
        "Fetching original ContractNotice lookup for item details: %s..%s. This is used to enrich awarded/completed notices.",
        date_from,
        date_to,
    )
    notices_by_tender_id: Dict[str, Dict[str, Any]] = {}
    scanned = 0
    kept = 0
    for window_from, window_to, page, notices in iter_notices(
        session,
        notice_type="ContractNotice",
        date_from=date_from,
        date_to=date_to,
        page_size=page_size,
        window_days=window_days,
        request_sleep=request_sleep,
    ):
        scanned += len(notices)
        for notice in notices:
            tender_id = one_line(notice.get("tenderId"))
            if not tender_id or tender_id in notices_by_tender_id:
                continue
            if not is_medical_notice(notice):
                continue
            notices_by_tender_id[tender_id] = notice
            kept += 1
        logging.info(
            "Original ContractNotice lookup page done: %s..%s page=%s | scanned=%s | kept medical originals=%s",
            window_from,
            window_to,
            page,
            scanned,
            kept,
        )
    save_original_contract_cache(cache_path, notices_by_tender_id, date_from, date_to)
    logging.info("Original ContractNotice lookup ready: %s medical originals kept", len(notices_by_tender_id))
    return notices_by_tender_id


def page_checkpoint_key(notice_type: str, window_from: str, window_to: str, page_number: int) -> str:
    return f"{notice_type}|{window_from}|{window_to}|page={page_number}"


def default_checkpoint_path(output: Path) -> Path:
    return output.with_suffix(output.suffix + ".checkpoint.json")


def load_checkpoint(path: Path) -> Dict[str, Any]:
    if not path.exists():
        return {"completed_pages": []}
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise SystemExit(f"Checkpoint file is not valid JSON: {path} ({exc})") from exc
    if not isinstance(payload, dict):
        raise SystemExit(f"Checkpoint file has an unexpected format: {path}")
    completed = payload.get("completed_pages")
    if not isinstance(completed, list):
        payload["completed_pages"] = []
    return payload


def save_checkpoint(path: Path, checkpoint: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    checkpoint["updated_at"] = utc_now_iso()
    temp_path = path.with_suffix(path.suffix + ".tmp")
    temp_path.write_text(json.dumps(checkpoint, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")
    temp_path.replace(path)


def validate_or_create_output(path: Path, fresh_start: bool = False) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if fresh_start and path.exists():
        path.unlink()
    if not path.exists() or path.stat().st_size == 0:
        with path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=FIELDNAMES, extrasaction="ignore")
            writer.writeheader()
        return

    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        existing_header = next(reader, [])
    if existing_header != FIELDNAMES:
        raise SystemExit(
            "Existing output CSV header does not match this script version. "
            "Use a new --output file or rerun with --fresh-start. "
            f"Problem file: {path}"
        )


def load_existing_dedup_keys(path: Path) -> set[str]:
    keys: set[str] = set()
    if not path.exists() or path.stat().st_size == 0:
        return keys
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            key = one_line(row.get("dedup_key"))
            if key and key != NA:
                keys.add(key)
    return keys


def append_csv_rows(path: Path, rows: Iterable[Dict[str, str]]) -> int:
    rows = list(rows)
    if not rows:
        return 0
    count = 0
    with path.open("a", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=FIELDNAMES, extrasaction="ignore")
        for row in rows:
            writer.writerow(fill_no_blanks_all(row))
            count += 1
        handle.flush()
        os.fsync(handle.fileno())
    return count


def progress_fraction(done: int, total: Optional[int]) -> str:
    if total is None:
        return f"{done}/unknown"
    return f"{done}/{total}"


def should_log_tender_progress(index: int, total: int, progress_every: int) -> bool:
    if total <= 0 or progress_every <= 0:
        return False
    return index == 1 or index == total or index % progress_every == 0


def build_progress_plan(
    session: requests.Session,
    notice_types: Sequence[str],
    date_from: str,
    date_to: str,
    page_size: int,
    window_days: int,
    request_sleep: float,
) -> Tuple[Dict[str, int], Optional[int], Optional[int]]:
    """Pre-count pages/notices so progress logs can show done/total.

    This makes one lightweight counting pass over the same API pages before the
    processing pass. The scraper still works if the count pass fails; progress
    then shows an unknown denominator.
    """
    page_notice_counts: Dict[str, int] = {}
    total_notices = 0
    total_pages = 0
    logging.info("Counting total API notices first so progress can show done/total...")
    try:
        for notice_type in notice_types:
            for window_from, window_to, page, notices in iter_notices(
                session,
                notice_type=notice_type,
                date_from=date_from,
                date_to=date_to,
                page_size=page_size,
                window_days=window_days,
                request_sleep=request_sleep,
            ):
                key = page_checkpoint_key(notice_type, window_from, window_to, page)
                count = len(notices)
                page_notice_counts[key] = count
                total_notices += count
                total_pages += 1
                logging.info(
                    "Counted page %s: %s tenders; running total %s tenders across %s pages",
                    f"live:{notice_type}:{window_from}..{window_to}:page={page}",
                    count,
                    total_notices,
                    total_pages,
                )
    except Exception as exc:
        logging.warning(
            "Could not pre-count total notices. The scraper will continue, but progress will show an unknown total. Error: %s",
            exc,
        )
        return {}, None, None

    logging.info("Total API tenders to scan: %s across %s pages", total_notices, total_pages)
    return page_notice_counts, total_notices, total_pages


def run_item_self_tests() -> int:
    samples = [
        (
            "table row",
            "1 Rękawice nitrylowe opak. 200",
            {"count": 1, "item_desc": "Rękawice nitrylowe", "item_uom": "opak", "item_quantity": "200"},
        ),
        (
            "lot sections",
            """### Część 1
### 4.2.2.) Krótki opis przedmiotu zamówienia
Implanty do zespoleń kości techniką blokowania kątowego
### 4.2.6.) Główny kod CPV: 33140000-3 - Materiały medyczne
### Część 2
### 4.2.2.) Krótki opis przedmiotu zamówienia
Waporyzator z akcesoriami
### 4.2.6.) Główny kod CPV: 33140000-3 - Materiały medyczne
""",
            {"count": 2, "first_desc_contains": "Implanty", "second_desc_contains": "Waporyzator"},
        ),
        (
            "reject contract period",
            "### 4.2.10.) Okres realizacji zamówienia albo umowy ramowej: 7 dni",
            {"count": 1, "item_uom": NA, "item_quantity": NA},
        ),
        (
            "reject hours as item quantity",
            "a) od momentu zgłoszenia przez Zamawiającego 24 godz.",
            {"count": 1, "item_uom": NA, "item_quantity": NA},
        ),
        (
            "remove supplier duplicate rows",
            """### Część 1
### 4.2.2.) Krótki opis przedmiotu zamówienia
Pakiet nr 1 – Aparat EKG
1 VERSAMED Sp. z o.o., Ul. Przędzalniana 14/1, 15-688 Białystok
""",
            {"count": 1, "first_desc_contains": "Aparat EKG"},
        ),
        (
            "reject delivery time as item",
            "Termin dostawy: od momentu zgłoszenia przez Zamawiającego - 24 godz.",
            {"count": 1, "item_uom": NA, "item_quantity": NA},
        ),
        (
            "spreadsheet tab row",
            "Lp\tNazwa asortymentu\tJ.m.\tIlość\n1\tStrzykawka 10 ml\tszt.\t500",
            {"count": 1, "item_desc": "Strzykawka 10 ml", "item_uom": "szt", "item_quantity": "500"},
        ),
        (
            "collapsed spreadsheet rows",
            "Lp. Nazwa asortymentu J.m. Ilość 1 Rękawice nitrylowe opak. 200 2 Strzykawka 10 ml szt. 500",
            {"count": 2, "item_desc": "Rękawice nitrylowe", "item_uom": "opak", "item_quantity": "200", "second_desc_contains": "Strzykawka"},
        ),
        (
            "do not treat dosage as quantity",
            "Strzykawka 10 ml",
            {"count": 1, "item_uom": NA, "item_quantity": NA},
        ),
    ]
    failures: List[str] = []
    for name, text, expected in samples:
        rows = extract_items(text, text, "Sample title")
        if len(rows) != expected.get("count"):
            failures.append(f"{name}: expected count {expected.get('count')}, got {len(rows)}: {rows}")
            continue
        first = rows[0]
        if "item_desc" in expected and first.get("item_desc") != expected["item_desc"]:
            failures.append(f"{name}: expected item_desc={expected['item_desc']!r}, got {first.get('item_desc')!r}")
        if "item_uom" in expected and first.get("item_uom") != expected["item_uom"]:
            failures.append(f"{name}: expected item_uom={expected['item_uom']!r}, got {first.get('item_uom')!r}")
        if "item_quantity" in expected and first.get("item_quantity") != expected["item_quantity"]:
            failures.append(f"{name}: expected item_quantity={expected['item_quantity']!r}, got {first.get('item_quantity')!r}")
        if "first_desc_contains" in expected and expected["first_desc_contains"] not in first.get("item_desc", ""):
            failures.append(f"{name}: first item does not contain {expected['first_desc_contains']!r}: {first}")
        if "second_desc_contains" in expected and len(rows) > 1 and expected["second_desc_contains"] not in rows[1].get("item_desc", ""):
            failures.append(f"{name}: second item does not contain {expected['second_desc_contains']!r}: {rows[1]}")
    if failures:
        print("Item self-tests failed:")
        for failure in failures:
            print(f"- {failure}")
        return 1
    print("Item self-tests passed.")
    return 0


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Scrape medical public procurement notices from the Polish eZamowienia BZP API and add English translation columns."
    )
    parser.add_argument("--date-from", default="2024-01-01", help="Start publication date, YYYY-MM-DD. Default: 2024-01-01")
    parser.add_argument("--date-to", default=date.today().isoformat(), help="End publication date, YYYY-MM-DD. Default: today")
    parser.add_argument(
        "--notice-types",
        nargs="+",
        default=DEFAULT_NOTICE_TYPES,
        choices=ALLOWED_NOTICE_TYPES,
        help="Notice types to scrape. Default: ContractNotice",
    )
    parser.add_argument("--page-size", type=int, default=100, help="API PageSize. The API example uses 100.")
    parser.add_argument("--window-days", type=int, default=31, help="Split long ranges into N-day windows before paginating.")
    parser.add_argument("--request-sleep", type=float, default=0.15, help="Seconds to sleep between API pages.")
    parser.add_argument("--translation-sleep", type=float, default=0.15, help="Seconds to sleep between translation calls.")
    parser.add_argument("--output", default="", help="Output CSV path. Default is generated from date range.")
    parser.add_argument(
        "--checkpoint",
        default="",
        help="Checkpoint JSON path for resume. Default: <output>.checkpoint.json",
    )
    parser.add_argument(
        "--original-cache",
        default="",
        help="Cache JSON for original ContractNotice lookup. Default: <output>.original_contracts.json",
    )
    parser.add_argument(
        "--original-lookback-days",
        type=int,
        default=ORIGINAL_NOTICE_LOOKBACK_DAYS,
        help="How many days before --date-from to search for original ContractNotice records used to enrich awarded/completed notices. Default: 730.",
    )
    parser.add_argument(
        "--disable-original-contract-fetch",
        action="store_true",
        help="Disable the original ContractNotice lookup. Faster, but item/lot details in awarded notices will be much weaker.",
    )
    parser.add_argument(
        "--refresh-original-cache",
        action="store_true",
        help="Rebuild the original ContractNotice cache even if it already exists.",
    )
    parser.add_argument(
        "--fetch-documents",
        dest="fetch_documents",
        action="store_true",
        default=True,
        help="Fetch public procedure documents/SWZ/OPZ/assortment attachments when the BZP notice itself lacks item quantities. Default: enabled.",
    )
    parser.add_argument(
        "--no-fetch-documents",
        dest="fetch_documents",
        action="store_false",
        help="Disable document fetching. Faster, but item quantities are often unavailable in awarded notices without attachments.",
    )
    parser.add_argument(
        "--use-playwright-documents",
        action="store_true",
        help="Use Playwright to render JavaScript-loaded e-Zamowienia procedure pages and discover more public document links. Use when item details are still missing.",
    )
    parser.add_argument(
        "--document-cache-dir",
        default="",
        help="Folder for downloaded procedure documents when --fetch-documents is used. Default: <output>.documents",
    )
    parser.add_argument(
        "--max-documents-per-tender",
        type=int,
        default=5,
        help="Maximum document attachments to inspect per tender when --fetch-documents is used. Default: 5.",
    )
    parser.add_argument(
        "--fresh-start",
        action="store_true",
        help="Delete the output CSV and checkpoint before starting. Use this only when you do not want to resume.",
    )
    parser.add_argument(
        "--contract-period-inclusive",
        action="store_true",
        help="Calculate contract period as inclusive days: end_date - start_date + 1. Default uses end_date - start_date.",
    )
    parser.add_argument("--all-notices", action="store_true", help="Do not filter to medical notices.")
    parser.add_argument("--no-translate", action="store_true", help="Create *_en columns but copy original values instead of calling deep-translator.")
    parser.add_argument(
        "--translate-fields",
        nargs="+",
        default=TRANSLATE_SOURCE_FIELDS,
        choices=TRANSLATE_SOURCE_FIELDS,
        help="Source fields to translate into *_en columns.",
    )
    parser.add_argument(
        "--max-items-per-notice",
        type=int,
        default=0,
        help="Maximum item rows to output per notice. Use 0 for unlimited. Default: 0.",
    )
    parser.add_argument(
        "--progress-every",
        type=int,
        default=10,
        help="Print tender-level progress every N notices on each page. Use 1 for every tender; use 0 to disable tender-level progress lines.",
    )
    parser.add_argument(
        "--skip-total-count",
        action="store_true",
        help="Do not pre-count all API pages. Startup is faster, but progress will show done/unknown instead of done/total.",
    )
    parser.add_argument(
        "--self-test-items",
        action="store_true",
        help="Run local item-extraction sample tests and exit. Does not call the live API.",
    )
    parser.add_argument("--log-level", default="INFO", choices=("DEBUG", "INFO", "WARNING", "ERROR"))
    return parser


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = build_arg_parser()
    args = parser.parse_args(argv)
    logging.basicConfig(level=getattr(logging, args.log_level), format="%(asctime)s %(levelname)s %(message)s")

    if args.self_test_items:
        return run_item_self_tests()

    try:
        parse_iso_date(args.date_from)
        parse_iso_date(args.date_to)
    except ValueError as exc:
        parser.error(f"Dates must use YYYY-MM-DD: {exc}")

    output = Path(args.output) if args.output else Path(f"poland_ezamowienia_medical_{args.date_from}_to_{args.date_to}.csv")
    checkpoint_path = Path(args.checkpoint) if args.checkpoint else default_checkpoint_path(output)
    original_cache_path = Path(args.original_cache) if args.original_cache else default_original_cache_path(output)
    document_cache_dir = Path(args.document_cache_dir) if args.document_cache_dir else Path(str(output) + ".documents")

    output_existed_before = output.exists()

    if args.fresh_start and checkpoint_path.exists():
        checkpoint_path.unlink()
    if args.fresh_start and args.refresh_original_cache and original_cache_path.exists():
        original_cache_path.unlink()

    validate_or_create_output(output, fresh_start=args.fresh_start)
    checkpoint = load_checkpoint(checkpoint_path)

    if not output_existed_before and checkpoint.get("completed_pages"):
        logging.warning(
            "Checkpoint exists but the output CSV did not exist before this run; ignoring completed page markers."
        )
        checkpoint["completed_pages"] = []

    completed_pages = set(str(key) for key in checkpoint.get("completed_pages", []))
    existing_dedup_keys = load_existing_dedup_keys(output)

    checkpoint.setdefault("script", "poland_ezamowienia_medical_scraper_resume_debug")
    checkpoint.setdefault("output", str(output.resolve()))
    checkpoint.setdefault("completed_pages", [])
    checkpoint["date_from"] = args.date_from
    checkpoint["date_to"] = args.date_to
    checkpoint["notice_types"] = list(args.notice_types)
    checkpoint["page_size"] = args.page_size
    checkpoint["window_days"] = args.window_days
    save_checkpoint(checkpoint_path, checkpoint)

    session = requests.Session()
    session.headers.update(
        {
            "User-Agent": "Mozilla/5.0 (compatible; PolandMedicalTenderScraper/2.1-resume-debug)",
            "Accept": "application/json,text/plain,*/*",
        }
    )

    if args.skip_total_count:
        page_notice_counts: Dict[str, int] = {}
        total_api_notices: Optional[int] = None
        total_api_pages: Optional[int] = None
        logging.info("Skipping total pre-count. Progress will show done/unknown for the full run.")
    else:
        page_notice_counts, total_api_notices, total_api_pages = build_progress_plan(
            session,
            notice_types=args.notice_types,
            date_from=args.date_from,
            date_to=args.date_to,
            page_size=args.page_size,
            window_days=args.window_days,
            request_sleep=args.request_sleep,
        )

    checkpoint["progress_total_api_notices"] = total_api_notices
    checkpoint["progress_total_api_pages"] = total_api_pages
    checkpoint.setdefault("completed_page_notice_counts", {})
    save_checkpoint(checkpoint_path, checkpoint)

    if completed_pages:
        logging.info(
            "Resume checkpoint contains %s completed API pages. Existing output has %s unique row keys.",
            len(completed_pages),
            len(existing_dedup_keys),
        )

    counted_completed_from_checkpoint = total_api_notices is not None
    if counted_completed_from_checkpoint:
        progress_seen_api = sum(page_notice_counts.get(key, 0) for key in completed_pages)
        progress_pages_done = sum(1 for key in completed_pages if key in page_notice_counts)
        if completed_pages:
            logging.info(
                "Resume progress already completed from checkpoint: API tenders %s; API pages %s",
                progress_fraction(progress_seen_api, total_api_notices),
                progress_fraction(progress_pages_done, total_api_pages),
            )
    else:
        progress_seen_api = 0
        progress_pages_done = 0

    translator = TranslationCache(enabled=not args.no_translate, sleep_seconds=args.translation_sleep)
    contract_notice_by_tender_id: Dict[str, Dict[str, Any]] = {}
    should_fetch_original_contracts = (
        not args.disable_original_contract_fetch
        and any(nt in {"TenderResultNotice", "ContractPerformingNotice"} for nt in args.notice_types)
    )
    if should_fetch_original_contracts:
        original_from = original_contract_from_date(args.date_from, args.original_lookback_days)
        contract_notice_by_tender_id = collect_original_contract_notice_map(
            session,
            date_from=original_from,
            date_to=args.date_to,
            page_size=args.page_size,
            window_days=args.window_days,
            request_sleep=args.request_sleep,
            cache_path=original_cache_path,
            refresh_cache=args.refresh_original_cache,
        )
        logging.info(
            "Original ContractNotice enrichment enabled: %s records available from %s to %s",
            len(contract_notice_by_tender_id),
            original_from,
            args.date_to,
        )
    else:
        logging.info("Original ContractNotice enrichment disabled or not needed for selected notice types.")

    total_seen = 0
    total_medical = 0
    total_written = 0
    skipped_non_medical = 0
    skipped_duplicates = 0
    document_items_cache: Dict[str, List[Dict[str, str]]] = {}

    for notice_type in args.notice_types:
        for window_from, window_to, page, notices in iter_notices(
            session,
            notice_type=notice_type,
            date_from=args.date_from,
            date_to=args.date_to,
            page_size=args.page_size,
            window_days=args.window_days,
            request_sleep=args.request_sleep,
        ):
            checkpoint_key = page_checkpoint_key(notice_type, window_from, window_to, page)
            query_text = f"live:{notice_type}:{window_from}..{window_to}:page={page}"
            page_total = len(notices)

            if checkpoint_key in completed_pages:
                if not counted_completed_from_checkpoint:
                    progress_seen_api += page_total
                    progress_pages_done += 1
                logging.info(
                    "Skipping completed page from checkpoint: %s | page tenders=%s | API tender progress=%s | API page progress=%s",
                    query_text,
                    page_total,
                    progress_fraction(progress_seen_api, total_api_notices),
                    progress_fraction(progress_pages_done, total_api_pages),
                )
                continue

            logging.info(
                "Starting %s | page tenders=%s | API tender progress before page=%s | API page progress=%s",
                query_text,
                page_total,
                progress_fraction(progress_seen_api, total_api_notices),
                progress_fraction(progress_pages_done, total_api_pages),
            )
            page_written = 0
            page_medical = 0
            page_skipped_non_medical = 0
            page_duplicate_rows = 0
            page_start_done = progress_seen_api

            try:
                for page_index, notice in enumerate(notices, start=1):
                    total_seen += 1
                    notice_tender_id = one_line(notice.get("tenderId"))
                    notice_number = one_line(notice.get("noticeNumber"))
                    notice_actual_type = one_line(notice.get("noticeType")) or notice_type
                    notice_label = notice_number or notice_tender_id or "N/A"
                    status_detail = "processed"

                    if notice_actual_type == "ContractNotice" and notice_tender_id and notice_tender_id not in contract_notice_by_tender_id:
                        contract_notice_by_tender_id[notice_tender_id] = notice

                    if not args.all_notices and not is_medical_notice(notice):
                        skipped_non_medical += 1
                        page_skipped_non_medical += 1
                        status_detail = "skipped non-medical"
                        if should_log_tender_progress(page_index, page_total, args.progress_every):
                            current_api_done = page_start_done + page_index
                            logging.info(
                                "Tender progress: API %s | page %s/%s | current API page %s | medical this run=%s | rows written this run=%s | skipped non-medical=%s | duplicate rows=%s | notice=%s | %s",
                                progress_fraction(current_api_done, total_api_notices),
                                page_index,
                                page_total,
                                progress_fraction(progress_pages_done + 1, total_api_pages),
                                total_medical,
                                total_written,
                                skipped_non_medical,
                                skipped_duplicates,
                                notice_label,
                                status_detail,
                            )
                        continue

                    total_medical += 1
                    page_medical += 1
                    original_notice = contract_notice_by_tender_id.get(notice_tender_id)
                    document_items: List[Dict[str, str]] = []
                    mapped_rows = map_notice_to_rows(
                        notice,
                        query_text=query_text,
                        original_notice=original_notice,
                        max_items_per_notice=args.max_items_per_notice,
                        contract_period_inclusive=args.contract_period_inclusive,
                        document_items=document_items,
                    )

                    # Award/result notices usually contain only lot names and refer to SWZ/OPZ/formularz
                    # attachments for quantities. Automatically inspect public documents only when the
                    # BZP notice/original ContractNotice did not produce item_quantity + item_uom.
                    if args.fetch_documents and notice_tender_id and mapped_rows_need_document_items(mapped_rows):
                        if notice_tender_id not in document_items_cache:
                            extra_pages = extract_page_urls_from_notice_text(
                                get_html_text(notice),
                                get_html_text(original_notice) if original_notice else "",
                            )
                            document_items_cache[notice_tender_id] = extract_items_from_procedure_documents(
                                session,
                                tender_id=notice_tender_id,
                                title=one_line(notice.get("orderObject")),
                                cache_dir=document_cache_dir,
                                max_documents=args.max_documents_per_tender,
                                max_items=args.max_items_per_notice,
                                extra_page_urls=extra_pages,
                                use_playwright=args.use_playwright_documents,
                            )
                        document_items = document_items_cache.get(notice_tender_id, [])
                        if document_items:
                            mapped_rows = map_notice_to_rows(
                                notice,
                                query_text=query_text,
                                original_notice=original_notice,
                                max_items_per_notice=args.max_items_per_notice,
                                contract_period_inclusive=args.contract_period_inclusive,
                                document_items=document_items,
                            )

                    rows_to_write: List[Dict[str, str]] = []
                    duplicate_rows_this_notice = 0
                    for row in mapped_rows:
                        translated_row = add_translation_columns(row, translator, args.translate_fields)
                        dedup_key = one_line(translated_row.get("dedup_key"))
                        if dedup_key and dedup_key in existing_dedup_keys:
                            skipped_duplicates += 1
                            page_duplicate_rows += 1
                            duplicate_rows_this_notice += 1
                            continue
                        if dedup_key:
                            existing_dedup_keys.add(dedup_key)
                        rows_to_write.append(translated_row)

                    written_now = append_csv_rows(output, rows_to_write)
                    page_written += written_now
                    total_written += written_now
                    status_detail = f"medical; rows written now={written_now}; duplicate rows now={duplicate_rows_this_notice}"

                    if should_log_tender_progress(page_index, page_total, args.progress_every):
                        current_api_done = page_start_done + page_index
                        logging.info(
                            "Tender progress: API %s | page %s/%s | current API page %s | medical this run=%s | rows written this run=%s | skipped non-medical=%s | duplicate rows=%s | notice=%s | %s",
                            progress_fraction(current_api_done, total_api_notices),
                            page_index,
                            page_total,
                            progress_fraction(progress_pages_done + 1, total_api_pages),
                            total_medical,
                            total_written,
                            skipped_non_medical,
                            skipped_duplicates,
                            notice_label,
                            status_detail,
                        )

            except Exception:
                logging.exception(
                    "Stopped while processing %s. Rows already appended remain in %s. "
                    "Rerun the same command to resume this page.",
                    query_text,
                    output,
                )
                raise

            progress_seen_api += page_total
            progress_pages_done += 1
            checkpoint["completed_pages"].append(checkpoint_key)
            checkpoint["completed_page_notice_counts"][checkpoint_key] = page_total
            checkpoint["last_completed_page"] = {
                "notice_type": notice_type,
                "window_from": window_from,
                "window_to": window_to,
                "page": page,
                "query_text": query_text,
                "api_tenders_on_page": page_total,
                "medical_tenders_on_page": page_medical,
                "skipped_non_medical_on_page": page_skipped_non_medical,
                "duplicate_rows_on_page": page_duplicate_rows,
                "rows_written_on_page": page_written,
            }
            checkpoint["totals"] = {
                "api_tenders_seen_this_run": total_seen,
                "medical_tenders_this_run": total_medical,
                "written_this_run": total_written,
                "skipped_non_medical_this_run": skipped_non_medical,
                "skipped_duplicates_this_run": skipped_duplicates,
                "api_tender_progress": progress_fraction(progress_seen_api, total_api_notices),
                "api_page_progress": progress_fraction(progress_pages_done, total_api_pages),
            }
            save_checkpoint(checkpoint_path, checkpoint)
            completed_pages.add(checkpoint_key)
            logging.info(
                "Completed %s | page tenders done=%s/%s | page medical=%s | page skipped non-medical=%s | page duplicate rows=%s | page rows written=%s | API tender progress=%s | API page progress=%s | total rows written this run=%s",
                query_text,
                page_total,
                page_total,
                page_medical,
                page_skipped_non_medical,
                page_duplicate_rows,
                page_written,
                progress_fraction(progress_seen_api, total_api_notices),
                progress_fraction(progress_pages_done, total_api_pages),
                total_written,
            )

    logging.info("Final API tender progress: %s", progress_fraction(progress_seen_api, total_api_notices))
    logging.info("Final API page progress: %s", progress_fraction(progress_pages_done, total_api_pages))
    logging.info("Processed %s API tenders this run", total_seen)
    logging.info("Matched %s medical tenders this run", total_medical)
    logging.info("Skipped %s non-medical notices", skipped_non_medical)
    logging.info("Skipped %s duplicate rows already present in output", skipped_duplicates)
    logging.info("Wrote %s new rows to %s", total_written, output)
    print(f"API tender progress: {progress_fraction(progress_seen_api, total_api_notices)}")
    print(f"API page progress: {progress_fraction(progress_pages_done, total_api_pages)}")
    print(f"API tenders processed this run: {total_seen}")
    print(f"Medical tenders matched this run: {total_medical}")
    print(f"Non-medical tenders skipped this run: {skipped_non_medical}")
    print(f"Duplicate rows skipped this run: {skipped_duplicates}")
    print(f"Rows written this run: {total_written}")
    print(f"Output: {output.resolve()}")
    print(f"Checkpoint: {checkpoint_path.resolve()}")
    if should_fetch_original_contracts:
        print(f"Original ContractNotice cache: {original_cache_path.resolve()}")
    if args.fetch_documents:
        print(f"Document cache dir: {document_cache_dir.resolve()}")
        if args.use_playwright_documents:
            print("Document discovery: requests + Playwright renderer")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
